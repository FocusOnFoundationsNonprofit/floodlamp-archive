"""
FloodLAMP public de-ID XLSX → curated CSV export + verification (v2).

This module curates a small set of high-signal, stable-schema CSVs from the published
de-identified (public) XLSX workbooks for a pilot site (APS / RFR / RTR).

Changes in v2 (relative to the initial draft):
- Treat any pandas-missing value (pd.isna) as missing in date/time normalization.
  This prevents literal "NaT"/"nan" strings from appearing in CSV outputs.
- Fix Stats referral agreement mini-table extraction to preserve legitimate zeros
  (avoid patterns like `value or ""` which turn 0 → "").
- Add explicit percent status columns where percent fields exist, so downstream
  users can distinguish:
    - ok        : percent is defined (denominator > 0)
    - denom_zero: percent undefined because denominator == 0
    - not_available: percent cannot be computed because required inputs are missing

NOTE:
- This module intentionally does NOT attempt to reproduce every sheet/tab.
- It focuses on the curated subset described in the project documentation.

"""

import csv
import datetime
import os
import re
import warnings

import numpy as np
import openpyxl
import pandas as pd


# -----------------------------
# Constants: paths
# -----------------------------

PILOT_DATA_ROOT = "data/floodlamp/pilots/pilot-data/"

# -----------------------------
# Constants: output filenames
# -----------------------------

def _out_name(site_code, file_type, name_slug):
    """Build a curated CSV filename with the standard naming convention.

    :param site_code: str, 4-letter site code (e.g., "ABRM").
    :param file_type: str, "APS" / "RFR" / "RTR".
    :param name_slug: str, filename tail (e.g., "weekly-summary").
    :return: str, filename like "ABRM_APS_weekly-summary.csv".
    """
    site_code = (site_code or "").strip().upper()
    file_type = (file_type or "").strip().upper()
    name_slug = (name_slug or "").strip()
    return f"{site_code}_{file_type}_{name_slug}.csv"


def _manifest_name(site_code):
    """Build a per-site manifest filename.

    :param site_code: str, site code.
    :return: str, filename like "ABRM_csv-manifest.csv".
    """
    site_code = (site_code or "").strip().upper()
    return f"{site_code}_csv-manifest.csv"


# -----------------------------
# Helpers: printing
# -----------------------------

def _print_section(title):
    """Print a section header.

    :param title: str, section title.
    :return: None
    """
    print(f"=== {title} ===")


def _print_kv(label, value):
    """Print a label/value line.

    :param label: str, label.
    :param value: any, value.
    :return: None
    """
    print(f"{label}: {value}")


def _print_dash():
    """Print a dashed divider.

    :return: None
    """
    print("---")


# -----------------------------
# Helpers: filesystem
# -----------------------------

def _ensure_folder(folder_path):
    """Ensure a folder exists.

    :param folder_path: str, folder path.
    :return: str, normalized folder path.
    """
    folder_path = os.path.abspath(folder_path)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


def _join(folder_path, filename):
    """Join folder + filename into a full path.

    :param folder_path: str, folder path.
    :param filename: str, filename.
    :return: str, full path.
    """
    return os.path.join(folder_path, filename)


# -----------------------------
# Helpers: normalization
# -----------------------------

def _norm_header(value):
    """Normalize a header/cell label to a clean string.

    :param value: any, header cell value.
    :return: str, normalized header (empty string if None).
    """
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _strip_trailing_total(header):
    """Strip embedded totals from a header like 'Weekly Participants 2954'.

    This is meant to stabilize schemas across sites.

    :param header: str, header text.
    :return: str, header with trailing numeric token removed (if present).
    """
    h = _norm_header(header)
    if not h:
        return h
    h = re.sub(r"\s+", " ", h).strip()
    parts = h.split(" ")
    if len(parts) >= 2 and re.fullmatch(r"[0-9]+(\.[0-9]+)?", parts[-1] or ""):
        return " ".join(parts[:-1]).strip()
    return h


def _is_missing(value):
    """Return True if a value should be treated as missing.

    This catches:
    - None
    - pandas missing sentinels (pd.NaT, pd.NA)
    - numpy NaN

    :param value: any
    :return: bool
    """
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except Exception:
        return False


def _to_iso_date(value):
    """Convert a date-like value to ISO YYYY-MM-DD, else return empty string.

    :param value: any, date/datetime-like or string.
    :return: str, ISO date or empty string.
    """
    if _is_missing(value):
        return ""

    if isinstance(value, pd.Timestamp):
        # pd.Timestamp is not pd.NaT (handled above), so safe to convert.
        value = value.to_pydatetime()

    if isinstance(value, datetime.datetime):
        return value.date().isoformat()

    if isinstance(value, datetime.date):
        return value.isoformat()

    # Some sheets store date as string already
    s = str(value).strip()
    if not s:
        return ""
    if s.lower() in {"nat", "nan"}:
        return ""

    # If it looks like "2022-09-05 00:00:00" keep the date part
    m = re.match(r"^(\d{4}-\d{2}-\d{2})(\s|$)", s)
    if m:
        return m.group(1)

    return s


def _to_hms_time(value):
    """Convert a time-like value to HH:MM:SS, else return empty string.

    :param value: any, time/datetime-like or string.
    :return: str, time string or empty string.
    """
    if _is_missing(value):
        return ""

    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()

    if isinstance(value, datetime.datetime):
        return value.time().strftime("%H:%M:%S")

    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")

    s = str(value).strip()
    if not s:
        return ""
    if s.lower() in {"nat", "nan"}:
        return ""

    # If it looks like "07:14:44" accept as-is
    m = re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", s)
    if m:
        if len(s.split(":")) == 2:
            s = f"{s}:00"
        parts = s.split(":")
        if len(parts[0]) == 1:
            s = f"0{s}"
        return s

    return s


def _to_iso_datetime(value):
    """Convert a datetime-like value to ISO YYYY-MM-DDTHH:MM:SS, else empty.

    :param value: any, datetime-like.
    :return: str, ISO datetime or empty string.
    """
    if _is_missing(value):
        return ""

    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()

    if isinstance(value, datetime.datetime):
        return value.replace(microsecond=0).isoformat()

    s = str(value).strip()
    if not s:
        return ""
    if s.lower() in {"nat", "nan"}:
        return ""

    return s


def _coerce_int(series):
    """Convert a pandas Series to nullable Int64.

    :param series: pandas.Series, input.
    :return: pandas.Series, Int64 series.
    """
    return pd.to_numeric(series, errors="coerce").astype("Int64")


def _coerce_float(series):
    """Convert a pandas Series to float.

    :param series: pandas.Series, input.
    :return: pandas.Series, float series.
    """
    return pd.to_numeric(series, errors="coerce").astype(float)


def _df_to_string_frame(df, float_format="%.15g"):
    """Convert a DataFrame into a DataFrame of strings matching CSV output.

    This is used for deterministic verification by comparing the expected
    content (from XLSX) to the actual CSV read back as strings.

    :param df: pandas.DataFrame, expected data.
    :param float_format: str, printf-style float format used when writing.
    :return: pandas.DataFrame, same shape with all values as strings.
    """
    def fmt_cell(v):
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass

        if isinstance(v, bool):
            return "True" if v else "False"

        if isinstance(v, (int,)):
            return str(v)

        # numpy numeric types
        try:
            if isinstance(v, (np.integer,)):
                return str(int(v))
            if isinstance(v, (np.floating,)):
                return float_format % float(v)
        except Exception:
            pass

        if isinstance(v, float):
            return float_format % v

        if isinstance(v, datetime.datetime):
            return v.replace(microsecond=0).isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        if isinstance(v, datetime.time):
            return v.strftime("%H:%M:%S")

        return str(v)

    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(fmt_cell)
    return out


def _pct_status(numerator, denominator):
    """Compute a percent status label for numerator/denominator.

    :param numerator: any, numerator value (may be missing).
    :param denominator: any, denominator value (may be missing).
    :return: str, one of: ok / denom_zero / not_available
    """
    if _is_missing(denominator):
        return "not_available"
    try:
        d = float(denominator)
    except Exception:
        return "not_available"

    if d == 0:
        return "denom_zero"
    return "ok"


def _finalize_pct(pct_value, numerator, denominator):
    """Finalize a percent value + status.

    Rules:
    - denom missing -> pct = NaN, status=not_available
    - denom == 0    -> pct = NaN, status=denom_zero
    - denom > 0:
        - if pct_value present -> use it
        - else if numerator present -> compute numerator/denominator
        - else -> pct = NaN, status=not_available

    :param pct_value: any, existing percent value (may be missing).
    :param numerator: any, numerator value (may be missing).
    :param denominator: any, denominator value (may be missing).
    :return: (float or np.nan, str status)
    """
    status = _pct_status(numerator, denominator)

    if status == "not_available":
        return (np.nan, status)

    try:
        d = float(denominator)
    except Exception:
        return (np.nan, "not_available")

    if d == 0:
        return (np.nan, "denom_zero")

    # denom > 0
    if not _is_missing(pct_value) and str(pct_value).strip() != "":
        try:
            return (float(pct_value), "ok")
        except Exception:
            # Fall back to compute if pct value isn't numeric
            pass

    if not _is_missing(numerator) and str(numerator).strip() != "":
        try:
            n = float(numerator)
            return (n / d, "ok")
        except Exception:
            return (np.nan, "not_available")

    return (np.nan, "not_available")




def _read_excel_sheet(xlsx_path, sheet_name, excel_file=None, header=0):
    """Read an Excel sheet into a DataFrame, reusing a pre-opened ExcelFile if provided.

    This is a performance optimization: loading a large XLSX multiple times can be slow.

    :param xlsx_path: str, path to XLSX (used if excel_file is None).
    :param sheet_name: str, worksheet/tab name.
    :param excel_file: pandas.ExcelFile or None, optional cached handle.
    :param header: int or None, header row for pandas.read_excel.
    :return: pandas.DataFrame
    """
    target = excel_file if excel_file is not None else xlsx_path
    return pd.read_excel(target, sheet_name=sheet_name, header=header)

# -----------------------------
# Helpers: CSV IO
# -----------------------------

def _write_csv(df, csv_path, float_format="%.15g"):
    """Write a DataFrame to CSV with deterministic settings.

    :param df: pandas.DataFrame, data to write.
    :param csv_path: str, output CSV path.
    :param float_format: str, float format string.
    :return: None
    """
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    df.to_csv(
        csv_path,
        index=False,
        encoding="utf-8",
        na_rep="",
        float_format=float_format,
        lineterminator="\n",
    )


def _read_csv_as_strings(csv_path):
    """Read a CSV into a DataFrame of strings (no NA coercion).

    :param csv_path: str, path to CSV.
    :return: pandas.DataFrame, all columns as strings; empty cells are "".
    """
    return pd.read_csv(csv_path, dtype=str, keep_default_na=False)


def _assert_frames_equal(expected_df, actual_df, context_label):
    """Assert two DataFrames (strings) are equal and raise with context.

    :param expected_df: pandas.DataFrame, expected values (strings).
    :param actual_df: pandas.DataFrame, actual values (strings).
    :param context_label: str, label for error context.
    :return: None
    """
    try:
        pd.testing.assert_frame_equal(expected_df, actual_df, check_dtype=False)
    except AssertionError as e:
        raise AssertionError(f"[VERIFY FAIL] {context_label}: {e}") from e


# -----------------------------
# APS builders
# -----------------------------

def _build_aps_stats_key_values_df(aps_xlsx_path, stats_ws=None):
    """Build the APS Stats key-values dataset from the Stats sheet.

    :param aps_xlsx_path: str, path to APS XLSX.
    :return: pandas.DataFrame, curated stats key-values.
    """
    if stats_ws is None:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True, keep_links=False)
        ws = wb["Stats"]
    else:
        ws = stats_ws
    def _norm_aux(v):
        if v is None:
            return ""
        if isinstance(v, datetime.datetime):
            return v.replace(microsecond=0).isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        if isinstance(v, datetime.time):
            return v.strftime("%H:%M:%S")
        return v

    current_section = "Files"
    records = []

    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value  # value
        b = ws.cell(row=r, column=2).value  # metric or section
        c = ws.cell(row=r, column=3).value  # details
        d = ws.cell(row=r, column=4).value  # comments

        b_norm = _norm_header(b)

        # Section headers: column C contains "Details" (capital D) or "section" (lowercase s).
        # Column B holds the section name when this is true.
        c_str = str(c) if c is not None else ""
        if "Details" in c_str or "section" in c_str:
            if b_norm:
                current_section = b_norm
            continue

        # Skip blank metric labels
        if b is None or (isinstance(b, str) and b.strip() == ""):
            continue

        # Include rows where metric name exists in B, even if value in A is blank.
        # Previously this excluded rows where a is None, but we want blank values preserved.
        value_norm = a if a is not None else ""
        if isinstance(a, (datetime.datetime, datetime.date)):
            value_norm = _to_iso_date(a)
        elif isinstance(a, datetime.time):
            value_norm = _to_hms_time(a)

        # Convert floats like 4.0 to int 4 for readability (but preserve non-integer floats)
        if isinstance(value_norm, float) and value_norm.is_integer():
            value_norm = int(value_norm)

        rec = {
            "section": current_section,
            "metric": b_norm,
            "value": value_norm,
            "details": c if c is not None else "",
            "comments": d if d is not None else "",
            "source_sheet": "Stats",
            "source_row": r,
        }
        records.append(rec)

    df = pd.DataFrame(
        records,
        columns=[
            "section",
            "metric",
            "value",
            "details",
            "comments",
            "source_sheet",
            "source_row",
        ],
    )
    return df


def _build_aps_stats_referral_agreement_df(aps_xlsx_path, stats_ws=None):
    """Build the APS Stats referral agreement mini-table dataset.

    v2 changes:
    - preserve legitimate zeros (no `value or ""`)
    - add percent status columns

    :param aps_xlsx_path: str, path to APS XLSX.
    :return: pandas.DataFrame, curated 2-row referral agreement table.
    """
    target = "Number of Tubes with FL Result"

    if stats_ws is None:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True, keep_links=False)
        ws = wb["Stats"]
    else:
        ws = stats_ws

    # Find anchor cell
    anchor_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == target:
                anchor_cell = cell
                break
        if anchor_cell is not None:
            break

    if anchor_cell is None:
        raise ValueError(f"Could not find anchor cell '{target}' in Stats sheet.")

    anchor_row = anchor_cell.row
    anchor_col = anchor_cell.column  # expected 6 (F)
    source_anchor = f"{openpyxl.utils.get_column_letter(anchor_col)}{anchor_row}"

    # Look for a breakdown header row that relabels I..L (9..12)
    breakdown_row = None
    for r in range(anchor_row + 1, min(anchor_row + 20, ws.max_row + 1)):
        v = _norm_header(ws.cell(row=r, column=9).value)  # I
        if v == "Ref/Cor Pos":
            breakdown_row = r
            break

    # Find data rows for Positive and Inconclusive (column E == 5)
    def find_category_row(category):
        for r in range(anchor_row + 1, ws.max_row + 1):
            v = _norm_header(ws.cell(row=r, column=5).value)
            if v == category:
                return r
        return None

    pos_row = find_category_row("Positive")
    inc_row = find_category_row("Inconclusive")

    if pos_row is None or inc_row is None:
        raise ValueError("Could not find both Positive and Inconclusive rows in Stats mini-table.")

    # Decide header mode for inconclusive:
    inc_use_breakdown = breakdown_row is not None and pos_row < breakdown_row < inc_row

    def blank_if_none(v):
        # Preserve 0; blank only if None
        return "" if v is None else v

    # Output columns (interleaving pct + pct_status for clarity)
    out_cols = [
        "fl_result_category",
        "tubes_n",
        "cases_n",
        "with_ref_or_corresp_n",
        "agree_n",
        "agree_pct",
        "agree_pct_status",
        "disagree_n",
        "disagree_pct",
        "disagree_pct_status",
        "ref_cor_pos_n",
        "incl_gt_pos_pct",
        "incl_gt_pos_pct_status",
        "ref_cor_neg_n",
        "incl_gt_neg_pct",
        "incl_gt_neg_pct_status",
        "source_sheet",
        "source_anchor",
    ]

    def read_row(category, row_num):
        tubes_n = ws.cell(row=row_num, column=6).value
        cases_n = ws.cell(row=row_num, column=7).value
        with_ref = ws.cell(row=row_num, column=8).value

        rec = {k: "" for k in out_cols}
        rec["fl_result_category"] = category
        rec["tubes_n"] = blank_if_none(tubes_n)
        rec["cases_n"] = blank_if_none(cases_n)
        rec["with_ref_or_corresp_n"] = blank_if_none(with_ref)

        # Read raw numeric / percent cells without losing 0
        if category == "Positive":
            agree_n = ws.cell(row=row_num, column=9).value
            agree_pct = ws.cell(row=row_num, column=10).value
            disagree_n = ws.cell(row=row_num, column=11).value
            disagree_pct = ws.cell(row=row_num, column=12).value

            rec["agree_n"] = blank_if_none(agree_n)
            rec["disagree_n"] = blank_if_none(disagree_n)

            pct_val, pct_status = _finalize_pct(agree_pct, agree_n, with_ref)
            rec["agree_pct"] = pct_val
            rec["agree_pct_status"] = pct_status

            pct_val, pct_status = _finalize_pct(disagree_pct, disagree_n, with_ref)
            rec["disagree_pct"] = pct_val
            rec["disagree_pct_status"] = pct_status

        else:
            # Inconclusive: breakdown if present, else fall back to main
            if inc_use_breakdown:
                ref_pos_n = ws.cell(row=row_num, column=9).value
                pct_pos = ws.cell(row=row_num, column=10).value
                ref_neg_n = ws.cell(row=row_num, column=11).value
                pct_neg = ws.cell(row=row_num, column=12).value

                rec["ref_cor_pos_n"] = blank_if_none(ref_pos_n)
                rec["ref_cor_neg_n"] = blank_if_none(ref_neg_n)

                pct_val, pct_status = _finalize_pct(pct_pos, ref_pos_n, with_ref)
                rec["incl_gt_pos_pct"] = pct_val
                rec["incl_gt_pos_pct_status"] = pct_status

                pct_val, pct_status = _finalize_pct(pct_neg, ref_neg_n, with_ref)
                rec["incl_gt_neg_pct"] = pct_val
                rec["incl_gt_neg_pct_status"] = pct_status

                # Agree/disagree are not the primary interpretation in breakdown mode,
                # but we still set a status so blanks are not ambiguous.
                rec["agree_pct_status"] = _pct_status(None, with_ref)
                rec["disagree_pct_status"] = _pct_status(None, with_ref)

            else:
                agree_n = ws.cell(row=row_num, column=9).value
                agree_pct = ws.cell(row=row_num, column=10).value
                disagree_n = ws.cell(row=row_num, column=11).value
                disagree_pct = ws.cell(row=row_num, column=12).value

                rec["agree_n"] = blank_if_none(agree_n)
                rec["disagree_n"] = blank_if_none(disagree_n)

                pct_val, pct_status = _finalize_pct(agree_pct, agree_n, with_ref)
                rec["agree_pct"] = pct_val
                rec["agree_pct_status"] = pct_status

                pct_val, pct_status = _finalize_pct(disagree_pct, disagree_n, with_ref)
                rec["disagree_pct"] = pct_val
                rec["disagree_pct_status"] = pct_status

                # Also set breakdown statuses based on denom, to clarify blanks
                rec["incl_gt_pos_pct_status"] = _pct_status(None, with_ref)
                rec["incl_gt_neg_pct_status"] = _pct_status(None, with_ref)

        rec["source_sheet"] = "Stats"
        rec["source_anchor"] = source_anchor
        return rec

    records = [
        read_row("Positive", pos_row),
        read_row("Inconclusive", inc_row),
    ]

    df = pd.DataFrame(records, columns=out_cols)

    # Normalize numeric-ish fields (counts as Int64; percents as float)
    int_cols = [
        "tubes_n",
        "cases_n",
        "with_ref_or_corresp_n",
        "agree_n",
        "disagree_n",
        "ref_cor_pos_n",
        "ref_cor_neg_n",
    ]
    float_cols = [
        "agree_pct",
        "disagree_pct",
        "incl_gt_pos_pct",
        "incl_gt_neg_pct",
    ]
    for col in int_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    for col in float_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype(float)

    return df



def _build_aps_stats_key_values_df_flmp(aps_xlsx_path, stats_ws=None, value_col=1):
    """Build APS Stats key-values for FLMP split Stats layout (FLSP/CRLN/Total columns).

    Layout: row 4 contains group headers; metric labels are in column D; details/comments in E/F.
    Values are in columns A (FLSP), B (CRLN), C (Total).

    :param aps_xlsx_path: str, APS XLSX path.
    :param stats_ws: openpyxl worksheet or None.
    :param value_col: int, 1=A(FLSP), 2=B(CRLN), 3=C(Total).
    :return: pandas.DataFrame, curated stats key-values.
    """
    if stats_ws is None:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True, keep_links=False)
        ws = wb["Stats"]
    else:
        ws = stats_ws

    def _norm_aux(v):
        if v is None:
            return ""
        if isinstance(v, datetime.datetime):
            return v.replace(microsecond=0).isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        if isinstance(v, datetime.time):
            return v.strftime("%H:%M:%S")
        return v

    current_section = "Files"
    records = []

    # File pointers (rows 1-3): use A=value, B=metric, C=details, D=comments
    for r in range(1, 4):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        b_norm = _norm_header(b)
        if not b_norm:
            continue
        value_norm = a
        if isinstance(a, (datetime.datetime, datetime.date)):
            value_norm = _to_iso_date(a)
        elif isinstance(a, datetime.time):
            value_norm = _to_hms_time(a)
        if isinstance(value_norm, float) and value_norm.is_integer():
            value_norm = int(value_norm)
        records.append({"section": current_section, "metric": b_norm, "value": value_norm if value_norm is not None else "", "details": _norm_aux(c), "comments": _norm_aux(d), "source_sheet": "Stats", "source_row": r})

    # Main Stats table begins at row 4 (section headers start here)
    for r, row in enumerate(ws.iter_rows(min_row=4, max_col=6, values_only=False), start=4):
        a1 = row[0].value
        a2 = row[1].value
        a3 = row[2].value
        metric = row[3].value
        details = row[4].value
        comments = row[5].value

        metric_norm = _norm_header(metric)

        # Section headers: column E (details) contains "Details" (capital D) or "section" (lowercase s).
        # Column D (metric) holds the section name when this is true.
        det_str = str(details) if details is not None else ""
        if "Details" in det_str or "section" in det_str:
            if metric_norm:
                current_section = metric_norm
            continue

        if not metric_norm:
            continue

        val = ws.cell(row=r, column=value_col).value
        value_norm = val
        if isinstance(val, (datetime.datetime, datetime.date)):
            value_norm = _to_iso_date(val)
        elif isinstance(val, datetime.time):
            value_norm = _to_hms_time(val)
        if isinstance(value_norm, float) and value_norm.is_integer():
            value_norm = int(value_norm)

        records.append({"section": current_section, "metric": metric_norm, "value": value_norm if value_norm is not None else "", "details": _norm_aux(details), "comments": _norm_aux(comments), "source_sheet": "Stats", "source_row": r})

    df = pd.DataFrame(records, columns=["section", "metric", "value", "details", "comments", "source_sheet", "source_row"])
    return df


def _build_aps_stats_referral_agreement_df_flmp(aps_xlsx_path, stats_ws=None, group_code="FLMP"):
    """Build APS Stats referral agreement mini-table for FLMP split Stats layout.

    The Stats sheet contains two mini-table blocks (FLSP and CRLN). This helper returns:
    - group_code="FLSP" -> 2-row table for FLSP
    - group_code="CRLN" -> 2-row table for CRLN
    - group_code="FLMP" -> 2-row combined table (sum of groups; pct recomputed)

    :param aps_xlsx_path: str, APS XLSX path.
    :param stats_ws: openpyxl worksheet or None.
    :param group_code: str, "FLSP" / "CRLN" / "FLMP".
    :return: pandas.DataFrame, curated 2-row referral agreement table.
    """
    target = "Number of Tubes with FL Result"
    if stats_ws is None:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True, keep_links=False)
        ws = wb["Stats"]
    else:
        ws = stats_ws

    anchor_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == target:
                anchor_cell = cell
                break
        if anchor_cell is not None:
            break
    if anchor_cell is None:
        raise ValueError(f"Could not find anchor cell '{target}' in Stats sheet.")

    anchor_row = anchor_cell.row
    anchor_col = anchor_cell.column  # expected I
    source_anchor = f"{openpyxl.utils.get_column_letter(anchor_col)}{anchor_row}"
    label_col = anchor_col - 1  # expected H

    flsp_pos = None
    crln_pos = None
    for r, row in enumerate(ws.iter_rows(min_row=1, min_col=1, max_col=4, values_only=False), start=1):
        if _norm_header(row[3].value) == "Number of Tubes with Final Result Positive":
            flsp_pos = row[0].value
            crln_pos = row[1].value
            break

    def _as_int(x):
        if _is_missing(x) or x is None:
            return None
        try:
            xf = float(x)
            if xf.is_integer():
                return int(xf)
            return xf
        except Exception:
            return x

    flsp_pos = _as_int(flsp_pos)
    crln_pos = _as_int(crln_pos)

    pos_rows = []
    for r, row in enumerate(ws.iter_rows(min_row=anchor_row + 1, max_row=anchor_row + 80, min_col=label_col, max_col=label_col, values_only=False), start=anchor_row + 1):
        if _norm_header(row[0].value) == "Positive":
            pos_rows.append(r)
    if not pos_rows:
        raise ValueError("Could not find any 'Positive' rows for FLMP referral agreement mini-table.")

    blocks = {}
    for pr in pos_rows:
        ir = None
        for r in range(pr + 1, pr + 21):
            if _norm_header(ws.cell(row=r, column=label_col).value) == "Inconclusive":
                ir = r
                break
        tubes_n = _as_int(ws.cell(row=pr, column=anchor_col).value)
        g = None
        if tubes_n is not None and flsp_pos is not None and tubes_n == flsp_pos:
            g = "FLSP"
        elif tubes_n is not None and crln_pos is not None and tubes_n == crln_pos:
            g = "CRLN"
        else:
            g = f"GROUP_{len(blocks)+1}"
        blocks[g] = (pr, ir)

    out_cols = [
        "fl_result_category",
        "tubes_n",
        "cases_n",
        "with_ref_or_corresp_n",
        "agree_n",
        "agree_pct",
        "agree_pct_status",
        "disagree_n",
        "disagree_pct",
        "disagree_pct_status",
        "ref_cor_pos_n",
        "incl_gt_pos_pct",
        "incl_gt_pos_pct_status",
        "ref_cor_neg_n",
        "incl_gt_neg_pct",
        "incl_gt_neg_pct_status",
        "source_sheet",
        "source_anchor",
    ]

    def read_positive_row(row_num):
        tubes_n = _as_int(ws.cell(row=row_num, column=anchor_col).value)
        cases_n = _as_int(ws.cell(row=row_num, column=anchor_col + 1).value)
        with_ref = _as_int(ws.cell(row=row_num, column=anchor_col + 2).value)
        agree_n = _as_int(ws.cell(row=row_num, column=anchor_col + 3).value)
        agree_pct_cell = ws.cell(row=row_num, column=anchor_col + 4).value
        disagree_n = _as_int(ws.cell(row=row_num, column=anchor_col + 5).value)
        disagree_pct_cell = ws.cell(row=row_num, column=anchor_col + 6).value
        agree_pct, agree_status = _finalize_pct(agree_pct_cell, agree_n, with_ref)
        disagree_pct, disagree_status = _finalize_pct(disagree_pct_cell, disagree_n, with_ref)
        return {
            "fl_result_category": "Positive",
            "tubes_n": tubes_n,
            "cases_n": cases_n,
            "with_ref_or_corresp_n": with_ref,
            "agree_n": agree_n,
            "agree_pct": agree_pct,
            "agree_pct_status": agree_status,
            "disagree_n": disagree_n,
            "disagree_pct": disagree_pct,
            "disagree_pct_status": disagree_status,
            "ref_cor_pos_n": "",
            "incl_gt_pos_pct": np.nan,
            "incl_gt_pos_pct_status": "not_available",
            "ref_cor_neg_n": "",
            "incl_gt_neg_pct": np.nan,
            "incl_gt_neg_pct_status": "not_available",
            "source_sheet": "Stats",
            "source_anchor": source_anchor,
        }

    def read_inconclusive_row(row_num):
        tubes_n = _as_int(ws.cell(row=row_num, column=anchor_col).value)
        cases_n = _as_int(ws.cell(row=row_num, column=anchor_col + 1).value)
        with_ref = _as_int(ws.cell(row=row_num, column=anchor_col + 2).value)
        ref_pos_n = _as_int(ws.cell(row=row_num, column=anchor_col + 3).value)
        incl_pos_pct_cell = ws.cell(row=row_num, column=anchor_col + 4).value
        ref_neg_n = _as_int(ws.cell(row=row_num, column=anchor_col + 5).value)
        incl_neg_pct_cell = ws.cell(row=row_num, column=anchor_col + 6).value
        incl_pos_pct, incl_pos_status = _finalize_pct(incl_pos_pct_cell, ref_pos_n, with_ref)
        incl_neg_pct, incl_neg_status = _finalize_pct(incl_neg_pct_cell, ref_neg_n, with_ref)
        return {
            "fl_result_category": "Inconclusive",
            "tubes_n": tubes_n,
            "cases_n": cases_n,
            "with_ref_or_corresp_n": with_ref,
            "agree_n": "",
            "agree_pct": np.nan,
            "agree_pct_status": "not_available",
            "disagree_n": "",
            "disagree_pct": np.nan,
            "disagree_pct_status": "not_available",
            "ref_cor_pos_n": ref_pos_n,
            "incl_gt_pos_pct": incl_pos_pct,
            "incl_gt_pos_pct_status": incl_pos_status,
            "ref_cor_neg_n": ref_neg_n,
            "incl_gt_neg_pct": incl_neg_pct,
            "incl_gt_neg_pct_status": incl_neg_status,
            "source_sheet": "Stats",
            "source_anchor": source_anchor,
        }

    def build_group(g):
        pr, ir = blocks.get(g, (None, None))
        if pr is None or ir is None:
            return None
        return pd.DataFrame([read_positive_row(pr), read_inconclusive_row(ir)], columns=out_cols)

    gnorm = (group_code or "").strip().upper()
    if gnorm in ("FLSP", "CRLN"):
        df = build_group(gnorm)
        if df is None:
            raise ValueError(f"Could not locate referral agreement block for group {group_code}.")
        return df

    df_flsp = build_group("FLSP")
    df_crln = build_group("CRLN")
    if df_flsp is None or df_crln is None:
        raise ValueError("Could not build both FLSP and CRLN referral agreement blocks for combined FLMP.")

    recs = []
    for cat in ["Positive", "Inconclusive"]:
        a = df_flsp[df_flsp["fl_result_category"] == cat].iloc[0].to_dict()
        b = df_crln[df_crln["fl_result_category"] == cat].iloc[0].to_dict()
        tubes_n = _as_int((0 if _is_missing(a["tubes_n"]) else a["tubes_n"]) + (0 if _is_missing(b["tubes_n"]) else b["tubes_n"]))
        cases_n = _as_int((0 if _is_missing(a["cases_n"]) else a["cases_n"]) + (0 if _is_missing(b["cases_n"]) else b["cases_n"]))
        with_ref = _as_int((0 if _is_missing(a["with_ref_or_corresp_n"]) else a["with_ref_or_corresp_n"]) + (0 if _is_missing(b["with_ref_or_corresp_n"]) else b["with_ref_or_corresp_n"]))
        if cat == "Positive":
            agree_n = _as_int((0 if _is_missing(a["agree_n"]) else a["agree_n"]) + (0 if _is_missing(b["agree_n"]) else b["agree_n"]))
            disagree_n = _as_int((0 if _is_missing(a["disagree_n"]) else a["disagree_n"]) + (0 if _is_missing(b["disagree_n"]) else b["disagree_n"]))
            agree_pct, agree_status = _finalize_pct(np.nan, agree_n, with_ref)
            disagree_pct, disagree_status = _finalize_pct(np.nan, disagree_n, with_ref)
            recs.append({
                "fl_result_category": cat,
                "tubes_n": tubes_n,
                "cases_n": cases_n,
                "with_ref_or_corresp_n": with_ref,
                "agree_n": agree_n,
                "agree_pct": agree_pct,
                "agree_pct_status": agree_status,
                "disagree_n": disagree_n,
                "disagree_pct": disagree_pct,
                "disagree_pct_status": disagree_status,
                "ref_cor_pos_n": "",
                "incl_gt_pos_pct": np.nan,
                "incl_gt_pos_pct_status": "not_available",
                "ref_cor_neg_n": "",
                "incl_gt_neg_pct": np.nan,
                "incl_gt_neg_pct_status": "not_available",
                "source_sheet": "Stats",
                "source_anchor": source_anchor,
            })
        else:
            ref_pos_n = _as_int((0 if _is_missing(a["ref_cor_pos_n"]) else a["ref_cor_pos_n"]) + (0 if _is_missing(b["ref_cor_pos_n"]) else b["ref_cor_pos_n"]))
            ref_neg_n = _as_int((0 if _is_missing(a["ref_cor_neg_n"]) else a["ref_cor_neg_n"]) + (0 if _is_missing(b["ref_cor_neg_n"]) else b["ref_cor_neg_n"]))
            incl_pos_pct, incl_pos_status = _finalize_pct(np.nan, ref_pos_n, with_ref)
            incl_neg_pct, incl_neg_status = _finalize_pct(np.nan, ref_neg_n, with_ref)
            recs.append({
                "fl_result_category": cat,
                "tubes_n": tubes_n,
                "cases_n": cases_n,
                "with_ref_or_corresp_n": with_ref,
                "agree_n": "",
                "agree_pct": np.nan,
                "agree_pct_status": "not_available",
                "disagree_n": "",
                "disagree_pct": np.nan,
                "disagree_pct_status": "not_available",
                "ref_cor_pos_n": ref_pos_n,
                "incl_gt_pos_pct": incl_pos_pct,
                "incl_gt_pos_pct_status": incl_pos_status,
                "ref_cor_neg_n": ref_neg_n,
                "incl_gt_neg_pct": incl_neg_pct,
                "incl_gt_neg_pct_status": incl_neg_status,
                "source_sheet": "Stats",
                "source_anchor": source_anchor,
            })

    df = pd.DataFrame(recs, columns=out_cols)
    return df


def _build_weekly_summary_df(xlsx_path, sheet_name, excel_file=None):
    """Build a weekly rollup dataset with stable headers.

    v2 changes:
    - add pct_positive_status
    - percent values are blanked when denominator==0 (status=denom_zero)

    :param xlsx_path: str, path to XLSX.
    :param sheet_name: str, sheet name.
    :return: pandas.DataFrame, curated weekly summary.
    """
    df_raw = _read_excel_sheet(xlsx_path, sheet_name=sheet_name, excel_file=excel_file)

    # Drop unnamed / empty columns
    cols = []
    for c in df_raw.columns:
        c_norm = str(c)
        if c_norm.startswith("Unnamed:"):
            continue
        cols.append(c)
    df_raw = df_raw[cols]

    # Identify needed columns by normalized header
    col_map = {}
    for c in df_raw.columns:
        c_norm = _norm_header(c).lower()
        if c_norm == "week start date":
            col_map["week_start_date"] = c
        elif c_norm == "week end date":
            col_map["week_end_date"] = c
        elif c_norm.startswith("weekly participants"):
            col_map["participants_n"] = c
        elif c_norm.startswith("weekly tubes"):
            # Handles "Weekly Tubes Run 1379" and "Weekly Tubes 1379"
            col_map["tubes_n"] = c
        elif c_norm.startswith("weekly positive tubes"):
            col_map["positive_tubes_n"] = c
        elif c_norm.startswith("weekly inconclusive tubes"):
            col_map["inconclusive_tubes_n"] = c
        elif "% positive" in c_norm:
            col_map["pct_positive"] = c

    required = [
        "week_start_date",
        "week_end_date",
        "participants_n",
        "tubes_n",
        "positive_tubes_n",
        "inconclusive_tubes_n",
        "pct_positive",
    ]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise ValueError(f"Missing weekly columns in {sheet_name}: {missing}")

    df = df_raw[[col_map[k] for k in required]].copy()
    df.columns = required

    # Drop blank rows
    df = df[df["week_start_date"].notna()].copy()

    # Normalize dates
    df["week_start_date"] = df["week_start_date"].map(_to_iso_date)
    df["week_end_date"] = df["week_end_date"].map(_to_iso_date)

    # Counts
    for c in ["participants_n", "tubes_n", "positive_tubes_n", "inconclusive_tubes_n"]:
        df[c] = _coerce_int(df[c])

    # pct_positive as float
    df["pct_positive"] = _coerce_float(df["pct_positive"])

    # Compute pct + status rowwise
    pct_vals = []
    pct_statuses = []
    for _, row in df.iterrows():
        pct_val, pct_status = _finalize_pct(
            pct_value=row["pct_positive"],
            numerator=row["positive_tubes_n"],
            denominator=row["tubes_n"],
        )
        pct_vals.append(pct_val)
        pct_statuses.append(pct_status)

    df["pct_positive"] = pct_vals
    df["pct_positive_status"] = pct_statuses

    # Deterministic ordering
    df = df.sort_values(["week_start_date", "week_end_date"], kind="mergesort").reset_index(drop=True)

    out_cols = [
        "week_start_date",
        "week_end_date",
        "participants_n",
        "tubes_n",
        "positive_tubes_n",
        "inconclusive_tubes_n",
        "pct_positive",
        "pct_positive_status",
    ]
    df = df[out_cols]
    return df


def _build_aps_pos_incl_cases_df(aps_xlsx_path, excel_file=None):
    """Build APS Pos/Inconclusive cases dataset.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated pos/incl tube-level dataset.
    """
    df_raw = _read_excel_sheet(aps_xlsx_path, sheet_name="Pos and Incl", excel_file=excel_file)
    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()

    accounted_col = "3 ACCOUNTED FOR IN RTR?"
    if accounted_col not in df_raw.columns:
        for c in df_raw.columns:
            if isinstance(c, str) and c.strip().upper().endswith("ACCOUNTED FOR IN RTR?"):
                accounted_col = c
                break

    pool_deconv_col = "Pool Deconvolution Indiv Test?"
    if pool_deconv_col not in df_raw.columns:
        if "Individual Deconvolution Test" in df_raw.columns:
            pool_deconv_col = "Individual Deconvolution Test"
        else:
            for c in df_raw.columns:
                if isinstance(c, str) and "DECONVOLUTION" in c.strip().upper() and "TEST" in c.strip().upper():
                    pool_deconv_col = c
                    break

    keep = {
        "TUBE ID": "tube_id",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "SPONSOR PERSON ID": "sponsor_id",
        "RESULT DATE": "result_date",
        "RESULT TIME": "result_time",
        "APP\nRESULT": "app_result",
        "FINAL RESULT": "final_result",
        "Result Correct Code": "result_correct_code",
        "Correction Notes": "correction_notes",
        "NUM SAMPLES (Pool Level)": "pool_level_n",
        "Sample Person1": "sample_person1",
        "Sample Person2": "sample_person2",
        "Sample Person3": "sample_person3",
        "Sample Person4": "sample_person4",
        "Case Culster ID": "case_cluster_id",
        "Index Tube - First FL Pos/Incl of Case?": "index_tube_first_case_flag",
        "Referral Test Result": "referral_test_result",
        "Referral Notes": "referral_notes",
        accounted_col: "accounted_for_in_rtr_flag",
        "Re-test?": "retest_flag",
        pool_deconv_col: "pool_deconvolution_indiv_test_flag",
        "Correspondence but no Referral Test": "correspondence_no_referral_flag",
        "Agreement with Referral or Correspondence?": "agreement_with_ref_or_correspondence_flag",
        "Known Positive (not Re-test)?": "known_positive_not_retest_flag",
        "FloodLAMP Result Eval": "floodlamp_result_eval",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    # Normalize date/time
    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["result_date"] = df["result_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["result_time"] = df["result_time"].map(_to_hms_time)

    # Coerce known numeric fields (keeps NA as <NA>)
    df["pool_level_n"] = _coerce_int(df["pool_level_n"])
    for c in [
        "case_cluster_id",
        "accounted_for_in_rtr_flag",
        "retest_flag",
        "pool_deconvolution_indiv_test_flag",
        "correspondence_no_referral_flag",
        "known_positive_not_retest_flag",
    ]:
        if c in df.columns:
            df[c] = _coerce_int(df[c])

    # Deterministic sort
    df = df.sort_values(["collection_date", "collection_time", "tube_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_aps_swabs_in_tubes_df(aps_xlsx_path, excel_file=None):
    """Build APS swabs-in-tubes participant-result dataset from postDEL.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated swab-in-tube dataset.
    """
    df_raw = _read_excel_sheet(aps_xlsx_path, sheet_name="postDEL", excel_file=excel_file)

    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()
    df_raw = df_raw[df_raw["Participant ID"].notna()].copy()

    keep = {
        "TUBE ID": "tube_id",
        "Participant ID": "participant_id",
        "Sponsor ID": "sponsor_id",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "RESULT DATE": "result_date",
        "RESULT TIME": "result_time",
        "APP\nRESULT": "app_result",
        "FINAL RESULT": "final_result",
        "APP\nGROUP": "app_group",
        "FINAL\nGROUP": "final_group",
        "Basic Correct Code": "basic_correct_code",
        "Result Correct Code": "result_correct_code",
        "Correction Notes": "correction_notes",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["result_date"] = df["result_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["result_time"] = df["result_time"].map(_to_hms_time)

    # IDs as strings (strip)
    for c in ["tube_id", "participant_id", "sponsor_id"]:
        df[c] = df[c].astype(str).str.strip()

    df = df.sort_values(["collection_date", "collection_time", "tube_id", "participant_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_aps_tubes_df(aps_xlsx_path, excel_file=None):
    """Build APS tube-level dataset from APS Tubes sheet.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated tubes dataset.
    """
    df_raw = _read_excel_sheet(aps_xlsx_path, sheet_name="APS Tubes", excel_file=excel_file)
    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()

    keep = {
        "TUBE ID": "tube_id",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "SPONSOR PERSON ID": "sponsor_id",
        "NUM SAMPLES (Pool Level)": "pool_level_n",
        "APP\nRESULT": "app_result",
        "FINAL RESULT": "final_result",
        "RESULT DATE": "result_date",
        "RESULT TIME": "result_time",
        "Result Correct Code": "result_correct_code",
        "Correction Notes": "correction_notes",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["result_date"] = df["result_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["result_time"] = df["result_time"].map(_to_hms_time)

    df["pool_level_n"] = _coerce_int(df["pool_level_n"])

    df = df.sort_values(["collection_date", "collection_time", "tube_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_aps_daily_tubes_rollup_df(aps_xlsx_path, excel_file=None):
    """Build APS daily tubes rollup dataset from APS Tubes by Date.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated daily rollup.
    """
    df_raw = _read_excel_sheet(aps_xlsx_path, sheet_name="APS Tubes by Date", excel_file=excel_file, header=None)

    # Slice leftmost A..I (0..8)
    df_slice = df_raw.iloc[:, 0:9].copy()

    header_row = df_slice.iloc[0].tolist()
    headers = [_strip_trailing_total(h) for h in header_row]
    headers = [_norm_header(h) for h in headers]

    df = df_slice.iloc[1:].copy()
    df.columns = headers

    first_col = headers[0]
    df = df[df[first_col].notna()].copy()

    rename_map = {
        "APS COLLECT DATE": "collect_date",
        "NUM TUBES": "tubes_n",
        "NEG": "neg_n",
        "POS": "pos_n",
        "INC": "inc_n",
        "NULL": "null_n",
        "INVL": "invl_n",
        "FAIL": "fail_n",
        "SUM RESULTS": "sum_results_n",
    }

    norm_to_original = {_norm_header(c): c for c in df.columns}
    rename_actual = {}
    for src_norm, dst in rename_map.items():
        if src_norm in norm_to_original:
            rename_actual[norm_to_original[src_norm]] = dst

    df = df.rename(columns=rename_actual).copy()

    required_cols = list(rename_map.values())
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"APS Tubes by Date missing expected columns after renaming: {missing}")

    df = df[required_cols].copy()

    df["collect_date"] = df["collect_date"].map(_to_iso_date)
    for c in ["tubes_n", "neg_n", "pos_n", "inc_n", "null_n", "invl_n", "fail_n", "sum_results_n"]:
        df[c] = _coerce_int(df[c])

    df = df.sort_values(["collect_date"], kind="mergesort").reset_index(drop=True)
    return df


# -----------------------------
# RFR builders
# -----------------------------

def _build_rfr_all_tube_results_df(rfr_xlsx_path, excel_file=None):
    """Build RFR all tube results dataset.

    :param rfr_xlsx_path: str, RFR XLSX path.
    :return: pandas.DataFrame, curated tube results with run linkage.
    """
    df_raw = _read_excel_sheet(rfr_xlsx_path, sheet_name="ALL Tube Results", excel_file=excel_file)
    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()

    keep = {
        "TUBE ID": "tube_id",
        "FINAL\nASSIGNED RUN ID": "final_assigned_run_id",
        "RUN DATE": "run_date",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "FINAL RESULT": "final_result",
        "NUM SAMPLES (Pool Level)": "pool_level_n",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    df["run_date"] = df["run_date"].map(_to_iso_date)
    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["pool_level_n"] = _coerce_int(df["pool_level_n"])

    df = df.sort_values(["run_date", "final_assigned_run_id", "tube_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


# -----------------------------
# RTR builders
# -----------------------------

def _build_rtr_referral_tests_df(rtr_xlsx_path, excel_file=None, sheet_name="Referral Test Data"):
    """Build RTR referral tests dataset from Referral Test Data.

    :param rtr_xlsx_path: str, RTR XLSX path.
    :return: pandas.DataFrame, curated referral tests (row-level).
    """
    df_raw = _read_excel_sheet(rtr_xlsx_path, sheet_name=sheet_name, excel_file=excel_file)
    df_raw = df_raw[df_raw["PARTICIPANT ID"].notna()].copy()

    keep = {
        "PARTICIPANT ID": "participant_id",
        "REFERRAL TEST COLLECTION DATE": "referral_test_collection_date",
        "REFERRAL TEST COLLECTION TIME": "referral_test_collection_time",
        "REFERRAL TEST RESULT DATE": "referral_test_result_date",
        "REFERRAL TEST RESULT TIME": "referral_test_result_time",
        "REFERRAL TEST TYPE": "referral_test_type",
        "REFERRAL TEST MODEL": "referral_test_model",
        "REFERRAL TEST RESULT": "referral_test_result",
        "REFERRAL TEST CT (IF PCR)": "referral_test_ct",
        "SYMPTOMS?": "symptoms_flag",
        "VACCINATED?": "vaccinated_flag",
        "SOURCE NOTES": "source_notes",
        "INIT FL POS TUBE ID FROM CASE CLUSTER": "init_fl_pos_tube_id",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    df["referral_test_collection_date"] = df["referral_test_collection_date"].map(_to_iso_date)
    df["referral_test_result_date"] = df["referral_test_result_date"].map(_to_iso_date)
    df["referral_test_collection_time"] = df["referral_test_collection_time"].map(_to_hms_time)
    df["referral_test_result_time"] = df["referral_test_result_time"].map(_to_hms_time)

    if "referral_test_ct" in df.columns:
        df["referral_test_ct"] = pd.to_numeric(df["referral_test_ct"], errors="coerce")

    df = df.sort_values(
        ["participant_id", "referral_test_collection_date", "referral_test_collection_time"],
        kind="mergesort",
    ).reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_rtr_referral_tests_by_person_df(rtr_xlsx_path, excel_file=None, sheet_name="Referral Tests by Person"):
    """Build RTR referral tests by person dataset.

    :param rtr_xlsx_path: str, RTR XLSX path.
    :return: pandas.DataFrame, curated per-person referral summary.
    """
    df_raw = _read_excel_sheet(rtr_xlsx_path, sheet_name=sheet_name, excel_file=excel_file)
    df_raw = df_raw[df_raw["PARTICIPANT ID"].notna()].copy()

    seq_col = "7 Num Sequential Referral Tests"
    if seq_col not in df_raw.columns:
        candidates = [c for c in df_raw.columns if isinstance(c, str) and c.strip().endswith("Num Sequential Referral Tests")]
        if candidates:
            seq_col = candidates[0]

    tube_id_col = "FloodLAMP Tube ID"
    if tube_id_col not in df_raw.columns:
        if "Corresponding FloodLAMP Tube ID" in df_raw.columns:
            tube_id_col = "Corresponding FloodLAMP Tube ID"
        else:
            candidates = [c for c in df_raw.columns if isinstance(c, str) and c.strip().endswith("FloodLAMP Tube ID")]
            if candidates:
                tube_id_col = candidates[0]

    keep = {
        "PARTICIPANT ID": "participant_id",
        seq_col: "num_sequential_referral_tests",
        "Num FloodLAMP Results Pos or Incl": "num_floodlamp_results_pos_or_incl",
        tube_id_col: "floodlamp_tube_id",
        "FloodLAMP Test Result": "floodlamp_test_result",
        "FloodLAMP Test Date": "floodlamp_test_date",
        "1st Referral Test Date": "first_referral_test_date",
        "Referral Overall Result": "referral_overall_result",
        "1st Referral Test Type": "first_referral_test_type",
        "1st Referral Test Result": "first_referral_test_result",
        "2nd Referral Test Type": "second_referral_test_type",
        "2nd Referral Test Result": "second_referral_test_result",
        "3rd Referral Test Type": "third_referral_test_type",
        "3rd Referral Test Result": "third_referral_test_result",
        "Antigen Neg with Other Positive": "antigen_neg_with_other_positive_flag",
        "Referral Eval": "referral_eval",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    df["floodlamp_test_date"] = df["floodlamp_test_date"].map(_to_iso_date)
    df["first_referral_test_date"] = df["first_referral_test_date"].map(_to_iso_date)

    df["num_sequential_referral_tests"] = _coerce_int(df["num_sequential_referral_tests"])
    df["num_floodlamp_results_pos_or_incl"] = _coerce_int(df["num_floodlamp_results_pos_or_incl"])

    df = df.sort_values(["participant_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


# -----------------------------
# Curate: write all outputs
# -----------------------------

def curate_site_public_csvs(aps_xlsx_path, rfr_xlsx_path, rtr_xlsx_path, output_folder_path, site_code):
    """Create all curated CSV outputs for a site from APS/RFR/RTR XLSX inputs.

    :param aps_xlsx_path: str, path to APS XLSX.
    :param rfr_xlsx_path: str, path to RFR XLSX.
    :param rtr_xlsx_path: str, path to RTR XLSX.
    :param output_folder_path: str, folder for CSV outputs.
    :param site_code: str, site code (e.g., ABRM).
    :return: dict, summary with written file paths and row counts.
    """
    output_folder_path = _ensure_folder(output_folder_path)
    site_code = (site_code or "").strip().upper()

    written = []
    rows_written = {}

    _print_section("CURATE: INPUTS")
    _print_kv("APS XLSX", aps_xlsx_path)
    _print_kv("RFR XLSX", rfr_xlsx_path)
    _print_kv("RTR XLSX", rtr_xlsx_path)
    _print_kv("Output folder", output_folder_path)
    _print_dash()

    # Performance: reuse parsed workbooks where possible
    aps_xl = pd.ExcelFile(aps_xlsx_path)
    rfr_xl = pd.ExcelFile(rfr_xlsx_path)
    rtr_xl = pd.ExcelFile(rtr_xlsx_path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        aps_wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True, keep_links=False, read_only=True)
    stats_ws = aps_wb["Stats"]

    # APS
    _print_section("CURATE: APS")
    aps_stats_kv_path = _join(output_folder_path, _out_name(site_code, "APS", "stats_key-values"))
    aps_stats_ref_path = _join(output_folder_path, _out_name(site_code, "APS", "stats_referral-agreement"))
    aps_weekly_path = _join(output_folder_path, _out_name(site_code, "APS", "weekly-summary"))
    aps_pos_incl_path = _join(output_folder_path, _out_name(site_code, "APS", "pos-incl-cases"))
    aps_swabs_path = _join(output_folder_path, _out_name(site_code, "APS", "swabs-in-tubes"))
    aps_tubes_path = _join(output_folder_path, _out_name(site_code, "APS", "tubes"))
    aps_daily_path = _join(output_folder_path, _out_name(site_code, "APS", "daily-tubes-rollup"))

    if site_code == "FLMP":
        # Helper to get group-specific output folder
        def _group_output_folder(g):
            if g in ("FLSP", "CRLN"):
                folder = os.path.join(PILOT_DATA_ROOT, f"{g}_pilot-data", f"{g}_curated_csvs")
                return _ensure_folder(folder)
            return output_folder_path
        # Clean up any existing CRLN/FLSP files from FLMP folder
        for g in ["FLSP", "CRLN"]:
            for file_type, slug in [("APS", "stats_key-values"), ("APS", "stats_referral-agreement"), ("RTR", "referral-tests"), ("RTR", "referral-tests-by-person")]:
                old_path = _join(output_folder_path, _out_name(g, file_type, slug))
                if os.path.exists(old_path):
                    os.remove(old_path)
                    _print_kv("Deleted from FLMP folder", os.path.basename(old_path))
        for g, col in [("FLSP", 1), ("CRLN", 2), ("FLMP", 3)]:
            g_folder = _group_output_folder(g)
            out_path = _join(g_folder, _out_name(g, "APS", "stats_key-values"))
            df_stats_kv = _build_aps_stats_key_values_df_flmp(aps_xlsx_path, stats_ws=stats_ws, value_col=col)
            _write_csv(df_stats_kv, out_path)
            written.append(out_path)
            rows_written[os.path.basename(out_path)] = len(df_stats_kv)
            _print_kv("Wrote", f"{out_path} ({len(df_stats_kv)} rows)")
        for g in ["FLSP", "CRLN", "FLMP"]:
            g_folder = _group_output_folder(g)
            out_path = _join(g_folder, _out_name(g, "APS", "stats_referral-agreement"))
            df_stats_ref = _build_aps_stats_referral_agreement_df_flmp(aps_xlsx_path, stats_ws=stats_ws, group_code=g)
            _write_csv(df_stats_ref, out_path)
            written.append(out_path)
            rows_written[os.path.basename(out_path)] = len(df_stats_ref)
            _print_kv("Wrote", f"{out_path} ({len(df_stats_ref)} rows)")
    else:
        df_stats_kv = _build_aps_stats_key_values_df(aps_xlsx_path, stats_ws=stats_ws)
        _write_csv(df_stats_kv, aps_stats_kv_path)
        written.append(aps_stats_kv_path)
        rows_written[os.path.basename(aps_stats_kv_path)] = len(df_stats_kv)
        _print_kv("Wrote", f"{os.path.basename(aps_stats_kv_path)} ({len(df_stats_kv)} rows)")

        df_stats_ref = _build_aps_stats_referral_agreement_df(aps_xlsx_path, stats_ws=stats_ws)
        _write_csv(df_stats_ref, aps_stats_ref_path)
        written.append(aps_stats_ref_path)
        rows_written[os.path.basename(aps_stats_ref_path)] = len(df_stats_ref)
        _print_kv("Wrote", f"{os.path.basename(aps_stats_ref_path)} ({len(df_stats_ref)} rows)")

    df_weekly = _build_weekly_summary_df(aps_xlsx_path, "REF Weekly", excel_file=aps_xl)
    _write_csv(df_weekly, aps_weekly_path)
    written.append(aps_weekly_path)
    rows_written[os.path.basename(aps_weekly_path)] = len(df_weekly)
    _print_kv("Wrote", f"{os.path.basename(aps_weekly_path)} ({len(df_weekly)} rows)")

    df_pos_incl = _build_aps_pos_incl_cases_df(aps_xlsx_path, excel_file=aps_xl)
    _write_csv(df_pos_incl, aps_pos_incl_path)
    written.append(aps_pos_incl_path)
    rows_written[os.path.basename(aps_pos_incl_path)] = len(df_pos_incl)
    _print_kv("Wrote", f"{os.path.basename(aps_pos_incl_path)} ({len(df_pos_incl)} rows)")

    df_swabs = _build_aps_swabs_in_tubes_df(aps_xlsx_path, excel_file=aps_xl)
    _write_csv(df_swabs, aps_swabs_path)
    written.append(aps_swabs_path)
    rows_written[os.path.basename(aps_swabs_path)] = len(df_swabs)
    _print_kv("Wrote", f"{os.path.basename(aps_swabs_path)} ({len(df_swabs)} rows)")

    df_tubes = _build_aps_tubes_df(aps_xlsx_path, excel_file=aps_xl)
    _write_csv(df_tubes, aps_tubes_path)
    written.append(aps_tubes_path)
    rows_written[os.path.basename(aps_tubes_path)] = len(df_tubes)
    _print_kv("Wrote", f"{os.path.basename(aps_tubes_path)} ({len(df_tubes)} rows)")

    df_daily = _build_aps_daily_tubes_rollup_df(aps_xlsx_path, excel_file=aps_xl)
    _write_csv(df_daily, aps_daily_path)
    written.append(aps_daily_path)
    rows_written[os.path.basename(aps_daily_path)] = len(df_daily)
    _print_kv("Wrote", f"{os.path.basename(aps_daily_path)} ({len(df_daily)} rows)")

    _print_dash()

    # RFR
    _print_section("CURATE: RFR")
    rfr_all_path = _join(output_folder_path, _out_name(site_code, "RFR", "all-tube-results"))
    rfr_weekly_path = _join(output_folder_path, _out_name(site_code, "RFR", "weekly-summary"))

    df_rfr_all = _build_rfr_all_tube_results_df(rfr_xlsx_path, excel_file=rfr_xl)
    _write_csv(df_rfr_all, rfr_all_path)
    written.append(rfr_all_path)
    rows_written[os.path.basename(rfr_all_path)] = len(df_rfr_all)
    _print_kv("Wrote", f"{os.path.basename(rfr_all_path)} ({len(df_rfr_all)} rows)")

    df_rfr_weekly = _build_weekly_summary_df(rfr_xlsx_path, "Weekly", excel_file=rfr_xl)
    _write_csv(df_rfr_weekly, rfr_weekly_path)
    written.append(rfr_weekly_path)
    rows_written[os.path.basename(rfr_weekly_path)] = len(df_rfr_weekly)
    _print_kv("Wrote", f"{os.path.basename(rfr_weekly_path)} ({len(df_rfr_weekly)} rows)")

    # FLMP: extract weekly summaries for CRLN and FLSP sub-programs from their dedicated tabs
    if site_code == "FLMP":
        for g in ["CRLN", "FLSP"]:
            g_folder = _group_output_folder(g)
            g_weekly_path = _join(g_folder, _out_name(g, "RFR", "weekly-summary"))
            df_g_weekly = _build_weekly_summary_df(rfr_xlsx_path, f"Weekly {g}", excel_file=rfr_xl)
            _write_csv(df_g_weekly, g_weekly_path)
            written.append(g_weekly_path)
            rows_written[os.path.basename(g_weekly_path)] = len(df_g_weekly)
            _print_kv("Wrote", f"{g_weekly_path} ({len(df_g_weekly)} rows)")

    _print_dash()

    # RTR
    _print_section("CURATE: RTR")
    if site_code == "FLMP":
        for g in ["CRLN", "FLSP"]:
            g_folder = _group_output_folder(g)
            tests_path = _join(g_folder, _out_name(g, "RTR", "referral-tests"))
            by_person_path = _join(g_folder, _out_name(g, "RTR", "referral-tests-by-person"))
            df_rtr_tests = _build_rtr_referral_tests_df(rtr_xlsx_path, excel_file=rtr_xl, sheet_name=f"Referral Test Data {g}")
            _write_csv(df_rtr_tests, tests_path)
            written.append(tests_path)
            rows_written[os.path.basename(tests_path)] = len(df_rtr_tests)
            _print_kv("Wrote", f"{tests_path} ({len(df_rtr_tests)} rows)")
            df_rtr_by_person = _build_rtr_referral_tests_by_person_df(rtr_xlsx_path, excel_file=rtr_xl, sheet_name=f"Referral Tests by Person {g}")
            _write_csv(df_rtr_by_person, by_person_path)
            written.append(by_person_path)
            rows_written[os.path.basename(by_person_path)] = len(df_rtr_by_person)
            _print_kv("Wrote", f"{by_person_path} ({len(df_rtr_by_person)} rows)")
    else:
        df_rtr_tests = _build_rtr_referral_tests_df(rtr_xlsx_path, excel_file=rtr_xl)
        _write_csv(df_rtr_tests, rtr_tests_path)
        written.append(rtr_tests_path)
        rows_written[os.path.basename(rtr_tests_path)] = len(df_rtr_tests)
        _print_kv("Wrote", f"{os.path.basename(rtr_tests_path)} ({len(df_rtr_tests)} rows)")

        df_rtr_by_person = _build_rtr_referral_tests_by_person_df(rtr_xlsx_path, excel_file=rtr_xl)
        _write_csv(df_rtr_by_person, rtr_by_person_path)
        written.append(rtr_by_person_path)
        rows_written[os.path.basename(rtr_by_person_path)] = len(df_rtr_by_person)
        _print_kv("Wrote", f"{os.path.basename(rtr_by_person_path)} ({len(df_rtr_by_person)} rows)")

    _print_dash()

    summary = {
        "written": written,
        "rows_written": rows_written,
    }

    _print_section("CURATE: SUMMARY")
    _print_kv("Files written", len(written))
    _print_kv("Row counts", rows_written)
    _print_section("CURATE: DONE")
    return summary


# -----------------------------
# Verify: compare CSVs to XLSX
# -----------------------------

def verify_site_public_csvs(aps_xlsx_path, rfr_xlsx_path, rtr_xlsx_path, output_folder_path, site_code):
    """Verify that all curated CSV outputs match the XLSX inputs under this module's rules.

    Verification strategy:
    - Recompute the expected curated DataFrame from the XLSX (using the same builders).
    - Read the produced CSV as strings (no NA coercion).
    - Convert expected DataFrame to strings using the same formatting rules.
    - Assert exact match in shape, columns, and cell values.

    :param aps_xlsx_path: str, APS XLSX path.
    :param rfr_xlsx_path: str, RFR XLSX path.
    :param rtr_xlsx_path: str, RTR XLSX path.
    :param output_folder_path: str, folder containing the curated CSVs.
    :param site_code: str, site code.
    :return: dict, verification summary (pass/fail per file).
    """
    output_folder_path = os.path.abspath(output_folder_path)
    site_code = (site_code or "").strip().upper()

    _print_section("VERIFY: INPUTS")
    _print_kv("APS XLSX", aps_xlsx_path)
    _print_kv("RFR XLSX", rfr_xlsx_path)
    _print_kv("RTR XLSX", rtr_xlsx_path)
    _print_kv("CSV folder", output_folder_path)
    _print_dash()

    # Performance: reuse parsed workbooks where possible
    aps_xl = pd.ExcelFile(aps_xlsx_path)
    rfr_xl = pd.ExcelFile(rfr_xlsx_path)
    rtr_xl = pd.ExcelFile(rtr_xlsx_path)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        aps_wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True, keep_links=False, read_only=True)
    stats_ws = aps_wb["Stats"]

    results = {}

    def verify_one(label, csv_filename, expected_df_builder):
        csv_path = _join(output_folder_path, csv_filename)
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"Missing expected CSV for verification: {csv_path}")

        expected_df = expected_df_builder()
        expected_str = _df_to_string_frame(expected_df)

        actual_str = _read_csv_as_strings(csv_path)

        if list(actual_str.columns) != list(expected_str.columns):
            raise AssertionError(
                f"[VERIFY FAIL] {label}: column mismatch\n"
                f"expected: {list(expected_str.columns)}\n"
                f"actual:   {list(actual_str.columns)}"
            )

        _assert_frames_equal(expected_str, actual_str, label)
        results[csv_filename] = "PASS"
        _print_kv("PASS", f"{label} → {csv_filename}")

    _print_section("VERIFY: APS")
    if site_code == "FLMP":
        # Helper to get group-specific output folder for verification
        def _group_verify_folder(g):
            if g in ("FLSP", "CRLN"):
                return os.path.join(PILOT_DATA_ROOT, f"{g}_pilot-data", f"{g}_curated_csvs")
            return output_folder_path
        for g, col in [("FLSP", 1), ("CRLN", 2), ("FLMP", 3)]:
            g_folder = _group_verify_folder(g)
            csv_path = _join(g_folder, _out_name(g, "APS", "stats_key-values"))
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f"Missing expected CSV for verification: {csv_path}")
            expected_df = _build_aps_stats_key_values_df_flmp(aps_xlsx_path, stats_ws=stats_ws, value_col=col)
            expected_str = _df_to_string_frame(expected_df)
            actual_str = _read_csv_as_strings(csv_path)
            label = f"APS Stats key-values ({g})"
            if list(actual_str.columns) != list(expected_str.columns):
                raise AssertionError(f"[VERIFY FAIL] {label}: column mismatch")
            _assert_frames_equal(expected_str, actual_str, label)
            results[os.path.basename(csv_path)] = "PASS"
            _print_kv("PASS", f"{label} → {csv_path}")
        for g in ["FLSP", "CRLN", "FLMP"]:
            g_folder = _group_verify_folder(g)
            csv_path = _join(g_folder, _out_name(g, "APS", "stats_referral-agreement"))
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f"Missing expected CSV for verification: {csv_path}")
            expected_df = _build_aps_stats_referral_agreement_df_flmp(aps_xlsx_path, stats_ws=stats_ws, group_code=g)
            expected_str = _df_to_string_frame(expected_df)
            actual_str = _read_csv_as_strings(csv_path)
            label = f"APS Stats referral agreement ({g})"
            if list(actual_str.columns) != list(expected_str.columns):
                raise AssertionError(f"[VERIFY FAIL] {label}: column mismatch")
            _assert_frames_equal(expected_str, actual_str, label)
            results[os.path.basename(csv_path)] = "PASS"
            _print_kv("PASS", f"{label} → {csv_path}")
    else:
        verify_one(
            "APS Stats key-values",
            _out_name(site_code, "APS", "stats_key-values"),
            lambda: _build_aps_stats_key_values_df(aps_xlsx_path, stats_ws=stats_ws),
        )
        verify_one(
            "APS Stats referral agreement",
            _out_name(site_code, "APS", "stats_referral-agreement"),
            lambda: _build_aps_stats_referral_agreement_df(aps_xlsx_path, stats_ws=stats_ws),
        )
    verify_one(
        "APS Weekly summary",
        _out_name(site_code, "APS", "weekly-summary"),
        lambda: _build_weekly_summary_df(aps_xlsx_path, "REF Weekly", excel_file=aps_xl),
    )
    verify_one(
        "APS Pos/Incl cases",
        _out_name(site_code, "APS", "pos-incl-cases"),
        lambda: _build_aps_pos_incl_cases_df(aps_xlsx_path, excel_file=aps_xl),
    )
    verify_one(
        "APS Swabs in tubes",
        _out_name(site_code, "APS", "swabs-in-tubes"),
        lambda: _build_aps_swabs_in_tubes_df(aps_xlsx_path, excel_file=aps_xl),
    )
    verify_one(
        "APS Tubes",
        _out_name(site_code, "APS", "tubes"),
        lambda: _build_aps_tubes_df(aps_xlsx_path, excel_file=aps_xl),
    )
    verify_one(
        "APS Daily tubes rollup",
        _out_name(site_code, "APS", "daily-tubes-rollup"),
        lambda: _build_aps_daily_tubes_rollup_df(aps_xlsx_path, excel_file=aps_xl),
    )
    _print_dash()

    _print_section("VERIFY: RFR")
    verify_one(
        "RFR All tube results",
        _out_name(site_code, "RFR", "all-tube-results"),
        lambda: _build_rfr_all_tube_results_df(rfr_xlsx_path, excel_file=rfr_xl),
    )
    verify_one(
        "RFR Weekly summary",
        _out_name(site_code, "RFR", "weekly-summary"),
        lambda: _build_weekly_summary_df(rfr_xlsx_path, "Weekly", excel_file=rfr_xl),
    )
    # FLMP: verify CRLN/FLSP weekly summaries
    if site_code == "FLMP":
        for g in ["CRLN", "FLSP"]:
            g_folder = _group_verify_folder(g)
            csv_path = _join(g_folder, _out_name(g, "RFR", "weekly-summary"))
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f"Missing expected CSV for verification: {csv_path}")
            expected_df = _build_weekly_summary_df(rfr_xlsx_path, f"Weekly {g}", excel_file=rfr_xl)
            expected_str = _df_to_string_frame(expected_df)
            actual_str = _read_csv_as_strings(csv_path)
            label = f"RFR Weekly summary ({g})"
            if list(actual_str.columns) != list(expected_str.columns):
                raise AssertionError(f"[VERIFY FAIL] {label}: column mismatch")
            _assert_frames_equal(expected_str, actual_str, label)
            results[os.path.basename(csv_path)] = "PASS"
            _print_kv("PASS", f"{label} → {csv_path}")
    _print_dash()

    _print_section("VERIFY: RTR")
    if site_code == "FLMP":
        for g in ["CRLN", "FLSP"]:
            g_folder = _group_verify_folder(g)
            # Referral tests
            csv_path = _join(g_folder, _out_name(g, "RTR", "referral-tests"))
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f"Missing expected CSV for verification: {csv_path}")
            expected_df = _build_rtr_referral_tests_df(rtr_xlsx_path, excel_file=rtr_xl, sheet_name=f"Referral Test Data {g}")
            expected_str = _df_to_string_frame(expected_df)
            actual_str = _read_csv_as_strings(csv_path)
            label = f"RTR Referral tests ({g})"
            if list(actual_str.columns) != list(expected_str.columns):
                raise AssertionError(f"[VERIFY FAIL] {label}: column mismatch")
            _assert_frames_equal(expected_str, actual_str, label)
            results[os.path.basename(csv_path)] = "PASS"
            _print_kv("PASS", f"{label} → {csv_path}")
            # Referral tests by person
            csv_path = _join(g_folder, _out_name(g, "RTR", "referral-tests-by-person"))
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f"Missing expected CSV for verification: {csv_path}")
            expected_df = _build_rtr_referral_tests_by_person_df(rtr_xlsx_path, excel_file=rtr_xl, sheet_name=f"Referral Tests by Person {g}")
            expected_str = _df_to_string_frame(expected_df)
            actual_str = _read_csv_as_strings(csv_path)
            label = f"RTR Referral tests by person ({g})"
            if list(actual_str.columns) != list(expected_str.columns):
                raise AssertionError(f"[VERIFY FAIL] {label}: column mismatch")
            _assert_frames_equal(expected_str, actual_str, label)
            results[os.path.basename(csv_path)] = "PASS"
            _print_kv("PASS", f"{label} → {csv_path}")
    else:
        verify_one(
            "RTR Referral tests",
            _out_name(site_code, "RTR", "referral-tests"),
            lambda: _build_rtr_referral_tests_df(rtr_xlsx_path, excel_file=rtr_xl),
        )
        verify_one(
            "RTR Referral tests by person",
            _out_name(site_code, "RTR", "referral-tests-by-person"),
            lambda: _build_rtr_referral_tests_by_person_df(rtr_xlsx_path, excel_file=rtr_xl),
        )
    _print_dash()

    _print_section("VERIFY: SUMMARY")
    _print_kv("Verified files", len(results))
    _print_kv("Results", results)
    _print_section("VERIFY: DONE")
    return results


# -----------------------------
# Manifest (optional convenience)
# -----------------------------

def write_site_csv_manifest(output_folder_path, site_code, aps_xlsx_name, rfr_xlsx_name, rtr_xlsx_name):
    """Write a per-site CSV manifest describing the curated CSV outputs.

    :param output_folder_path: str, output folder.
    :param site_code: str, site code.
    :param aps_xlsx_name: str, APS workbook filename.
    :param rfr_xlsx_name: str, RFR workbook filename.
    :param rtr_xlsx_name: str, RTR workbook filename.
    :return: str, path to written manifest CSV.
    """
    output_folder_path = _ensure_folder(output_folder_path)
    site_code = (site_code or "").strip().upper()

    manifest_path = _join(output_folder_path, _manifest_name(site_code))

    rows = [
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "Stats",
            "output_csv": _out_name(site_code, "APS", "stats_key-values"),
            "data_level": "stats",
            "primary_key": "",
            "join_keys": "",
            "row_grain": "one row per metric",
            "notes": "Tidy key/value extraction from Stats A-D; sectioned by header rows in col B.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "Stats",
            "output_csv": _out_name(site_code, "APS", "stats_referral-agreement"),
            "data_level": "stats",
            "primary_key": "",
            "join_keys": "",
            "row_grain": "one row per FL result category",
            "notes": "Extract mini-table anchored at 'Number of Tubes with FL Result' (cols E-L); includes pct status columns.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "REF Weekly",
            "output_csv": _out_name(site_code, "APS", "weekly-summary"),
            "data_level": "weekly_rollup",
            "primary_key": "",
            "join_keys": "week_start_date;week_end_date",
            "row_grain": "one row per week",
            "notes": "Weekly screening rollup (imported from RFR when present); includes pct_positive_status.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "Pos and Incl",
            "output_csv": _out_name(site_code, "APS", "pos-incl-cases"),
            "data_level": "tube",
            "primary_key": "tube_id",
            "join_keys": "tube_id;case_cluster_id",
            "row_grain": "one row per pos/incl tube",
            "notes": "Case-centric pos/incl subset with pool membership + referral linkage.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "postDEL",
            "output_csv": _out_name(site_code, "APS", "swabs-in-tubes"),
            "data_level": "participant_result",
            "primary_key": "",
            "join_keys": "tube_id;participant_id",
            "row_grain": "one row per swab-in-tube row",
            "notes": "Trimmed participant-result table; drops QA scaffolding and blank PII columns.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "APS Tubes",
            "output_csv": _out_name(site_code, "APS", "tubes"),
            "data_level": "tube",
            "primary_key": "tube_id",
            "join_keys": "tube_id",
            "row_grain": "one row per tube",
            "notes": "Canonical tube table (Appivo tubes); pool level + timing + results + corrections.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "APS Tubes by Date",
            "output_csv": _out_name(site_code, "APS", "daily-tubes-rollup"),
            "data_level": "daily_rollup",
            "primary_key": "collect_date",
            "join_keys": "collect_date",
            "row_grain": "one row per collection date",
            "notes": "Exports only leftmost A-I rollup table; headers normalized to remove embedded totals.",
        },
        {
            "site_code": site_code,
            "source_workbook": rfr_xlsx_name,
            "source_tab": "ALL Tube Results",
            "output_csv": _out_name(site_code, "RFR", "all-tube-results"),
            "data_level": "tube",
            "primary_key": "tube_id",
            "join_keys": "tube_id;final_assigned_run_id",
            "row_grain": "one row per tube",
            "notes": "Run linkage + includes ARF tubes (not in Appivo).",
        },
        {
            "site_code": site_code,
            "source_workbook": rfr_xlsx_name,
            "source_tab": "Weekly",
            "output_csv": _out_name(site_code, "RFR", "weekly-summary"),
            "data_level": "weekly_rollup",
            "primary_key": "",
            "join_keys": "week_start_date;week_end_date",
            "row_grain": "one row per week",
            "notes": "Canonical RFR weekly rollup; includes pct_positive_status.",
        },
        {
            "site_code": site_code,
            "source_workbook": rtr_xlsx_name,
            "source_tab": "Referral Test Data",
            "output_csv": _out_name(site_code, "RTR", "referral-tests"),
            "data_level": "referral",
            "primary_key": "referral_test_id",
            "join_keys": "participant_id;case_cluster_id;init_fl_pos_tube_id",
            "row_grain": "one row per referral test",
            "notes": "Trimmed referral test rows; keeps SOURCE NOTES; normalized missing times to blank.",
        },
        {
            "site_code": site_code,
            "source_workbook": rtr_xlsx_name,
            "source_tab": "Referral Tests by Person",
            "output_csv": _out_name(site_code, "RTR", "referral-tests-by-person"),
            "data_level": "referral",
            "primary_key": "participant_id",
            "join_keys": "participant_id",
            "row_grain": "one row per participant",
            "notes": "Per-person referral summary; drops PII and internal matching helper columns.",
        },
    ]

    fieldnames = [
        "site_code",
        "source_workbook",
        "source_tab",
        "output_csv",
        "data_level",
        "primary_key",
        "join_keys",
        "row_grain",
        "notes",
    ]

    with open(manifest_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    _print_section("MANIFEST")
    _print_kv("Wrote", manifest_path)
    _print_kv("Rows", len(rows))
    return manifest_path


# -----------------------------
# Manual run pattern
# -----------------------------

def mrun_curate_and_verify():
    """Manual run stub (fill in paths, then run).

    :return: None
    """
    pass


if __name__ == "__main__":
   site_code = "FLMP"
   site_folder = "data/floodlamp/pilots/pilot-data/" + site_code + "_pilot-data"
   downloads_folder = site_folder + "/" + site_code + "_xlsx_downloads"
   aps_xlsx_path = os.path.join(downloads_folder, site_code + "_APS_deID_PUB.xlsx")
   rfr_xlsx_path = os.path.join(downloads_folder, site_code + "_RFR_deID_PUB.xlsx")
   rtr_xlsx_path = os.path.join(downloads_folder, site_code + "_RTR_deID_PUB.xlsx")
   output_folder_path = site_folder + "/" + site_code + "_curated_csvs"

   curate_site_public_csvs(
       aps_xlsx_path=aps_xlsx_path,
       rfr_xlsx_path=rfr_xlsx_path,
       rtr_xlsx_path=rtr_xlsx_path,
       output_folder_path=output_folder_path,
       site_code=site_code,
   )

   verify_site_public_csvs(
       aps_xlsx_path=aps_xlsx_path,
       rfr_xlsx_path=rfr_xlsx_path,
       rtr_xlsx_path=rtr_xlsx_path,
       output_folder_path=output_folder_path,
       site_code=site_code,
   )

   write_site_csv_manifest(
       output_folder_path=output_folder_path,
       site_code=site_code,
       aps_xlsx_name=os.path.basename(aps_xlsx_path),
       rfr_xlsx_name=os.path.basename(rfr_xlsx_path),
       rtr_xlsx_name=os.path.basename(rtr_xlsx_path),
   )
