"""
FloodLAMP public de-ID XLSX → curated CSV export + verification.

This module curates a small set of high-signal, stable-schema CSVs from the published
de-identified (public) XLSX workbooks for a pilot site (APS / RFR / RTR).

The curated CSVs are intended for:
- human-readable archive consumption (paired with separately generated Markdown)
- downstream scientific analysis using a stable schema across sites

NOTE:
- This module intentionally does NOT attempt to reproduce every sheet/tab.
- It focuses on the curated subset described in the project documentation.

"""

import csv
import datetime
import os
import re
import warnings

import openpyxl
import pandas as pd


# -----------------------------
# Constants: output filenames
# -----------------------------

def _out_name(site_code, file_type, name_slug):
    """Build a curated CSV filename with the standard naming convention.

    :param site_code: str, 4-letter site code (e.g., "ABRM").
    :param file_type: str, "APS" / "RFR" / "RTR".
    :param name_slug: str, filename tail (e.g., "weekly-summary").
    :return: str, filename like "ABRM_APS_weekly-summary.csv".
    """
    site_code = (site_code or "").strip().upper()
    file_type = (file_type or "").strip().upper()
    name_slug = (name_slug or "").strip()
    return f"{site_code}_{file_type}_{name_slug}.csv"


def _manifest_name(site_code):
    """Build a per-site manifest filename.

    :param site_code: str, site code.
    :return: str, filename like "ABRM_csv-manifest.csv".
    """
    site_code = (site_code or "").strip().upper()
    return f"{site_code}_csv-manifest.csv"


# -----------------------------
# Helpers: printing
# -----------------------------

def _print_section(title):
    """Print a section header.

    :param title: str, section title.
    :return: None
    """
    print(f"=== {title} ===")


def _print_kv(label, value):
    """Print a label/value line.

    :param label: str, label.
    :param value: any, value.
    :return: None
    """
    print(f"{label}: {value}")


def _print_dash():
    """Print a dashed divider.

    :return: None
    """
    print("---")


# -----------------------------
# Helpers: filesystem
# -----------------------------

def _ensure_folder(folder_path):
    """Ensure a folder exists.

    :param folder_path: str, folder path.
    :return: str, normalized folder path.
    """
    folder_path = os.path.abspath(folder_path)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


def _join(folder_path, filename):
    """Join folder + filename into a full path.

    :param folder_path: str, folder path.
    :param filename: str, filename.
    :return: str, full path.
    """
    return os.path.join(folder_path, filename)


# -----------------------------
# Helpers: normalization
# -----------------------------

def _norm_header(value):
    """Normalize a header/cell label to a clean string.

    :param value: any, header cell value.
    :return: str, normalized header (empty string if None).
    """
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _strip_trailing_total(header):
    """Strip embedded totals from a header like 'Weekly Participants 2954'.

    This is meant to stabilize schemas across sites.

    :param header: str, header text.
    :return: str, header with trailing numeric token removed (if present).
    """
    h = _norm_header(header)
    if not h:
        return h
    # Collapse "POS  8" → "POS 8"
    h = re.sub(r"\s+", " ", h).strip()
    # If last token is numeric, drop it.
    parts = h.split(" ")
    if len(parts) >= 2 and re.fullmatch(r"[0-9]+(\.[0-9]+)?", parts[-1] or ""):
        return " ".join(parts[:-1]).strip()
    return h


def _to_iso_date(value):
    """Convert a date-like value to ISO YYYY-MM-DD, else return empty string.

    :param value: any, date/datetime-like or string.
    :return: str, ISO date or empty string.
    """
    if value is None:
        return ""
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        value = value.to_pydatetime()
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    # Some sheets store date as string already
    s = str(value).strip()
    if not s:
        return ""
    # If it looks like "2022-09-05 00:00:00" keep the date part
    m = re.match(r"^(\d{4}-\d{2}-\d{2})(\s|$)", s)
    if m:
        return m.group(1)
    return s


def _to_hms_time(value):
    """Convert a time-like value to HH:MM:SS, else return empty string.

    :param value: any, time/datetime-like or string.
    :return: str, time string or empty string.
    """
    if value is None:
        return ""
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        value = value.to_pydatetime()
    if isinstance(value, datetime.datetime):
        return value.time().strftime("%H:%M:%S")
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M:%S")
    s = str(value).strip()
    if not s:
        return ""
    # If it looks like "07:14:44" accept as-is
    m = re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", s)
    if m:
        if len(s.split(":")) == 2:
            return f"{s}:00"
        # Pad hour if needed
        parts = s.split(":")
        if len(parts[0]) == 1:
            return f"0{s}"
        return s
    return s


def _to_iso_datetime(value):
    """Convert a datetime-like value to ISO YYYY-MM-DDTHH:MM:SS, else empty.

    :param value: any, datetime-like.
    :return: str, ISO datetime or empty string.
    """
    if value is None:
        return ""
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        value = value.to_pydatetime()
    if isinstance(value, datetime.datetime):
        return value.replace(microsecond=0).isoformat()
    s = str(value).strip()
    return s


def _coerce_int(series):
    """Convert a pandas Series to nullable Int64.

    :param series: pandas.Series, input.
    :return: pandas.Series, Int64 series.
    """
    return pd.to_numeric(series, errors="coerce").astype("Int64")


def _coerce_float(series):
    """Convert a pandas Series to float.

    :param series: pandas.Series, input.
    :return: pandas.Series, float series.
    """
    return pd.to_numeric(series, errors="coerce").astype(float)


def _df_to_string_frame(df, float_format="%.15g"):
    """Convert a DataFrame into a DataFrame of strings matching CSV output.

    This is used for deterministic verification by comparing the expected
    content (from XLSX) to the actual CSV read back as strings.

    :param df: pandas.DataFrame, expected data.
    :param float_format: str, printf-style float format used when writing.
    :return: pandas.DataFrame, same shape with all values as strings.
    """
    def fmt_cell(v):
        if v is None:
            return ""
        # pandas NA / NaN
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass
        if isinstance(v, bool):
            return "True" if v else "False"
        if isinstance(v, (int,)):
            return str(v)
        # numpy numeric types
        if hasattr(v, "dtype"):
            # numpy scalar
            try:
                import numpy as np
                if isinstance(v, (np.integer,)):
                    return str(int(v))
                if isinstance(v, (np.floating,)):
                    return float_format % float(v)
            except Exception:
                pass
        if isinstance(v, float):
            return float_format % v
        if isinstance(v, datetime.datetime):
            return v.replace(microsecond=0).isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        if isinstance(v, datetime.time):
            return v.strftime("%H:%M:%S")
        return str(v)

    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(fmt_cell)
    return out


# -----------------------------
# Helpers: CSV IO
# -----------------------------

def _write_csv(df, csv_path, float_format="%.15g"):
    """Write a DataFrame to CSV with deterministic settings.

    :param df: pandas.DataFrame, data to write.
    :param csv_path: str, output CSV path.
    :param float_format: str, float format string.
    :return: None
    """
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    df.to_csv(
        csv_path,
        index=False,
        encoding="utf-8",
        na_rep="",
        float_format=float_format,
        lineterminator="\n",
    )


def _read_csv_as_strings(csv_path):
    """Read a CSV into a DataFrame of strings (no NA coercion).

    :param csv_path: str, path to CSV.
    :return: pandas.DataFrame, all columns as strings; empty cells are "".
    """
    return pd.read_csv(csv_path, dtype=str, keep_default_na=False)


def _assert_frames_equal(expected_df, actual_df, context_label):
    """Assert two DataFrames (strings) are equal and raise with context.

    :param expected_df: pandas.DataFrame, expected values (strings).
    :param actual_df: pandas.DataFrame, actual values (strings).
    :param context_label: str, label for error context.
    :return: None
    """
    try:
        pd.testing.assert_frame_equal(expected_df, actual_df, check_dtype=False)
    except AssertionError as e:
        raise AssertionError(f"[VERIFY FAIL] {context_label}: {e}") from e


# -----------------------------
# APS builders
# -----------------------------

def _build_aps_stats_key_values_df(aps_xlsx_path):
    """Build the APS Stats key-values dataset from the Stats sheet.

    :param aps_xlsx_path: str, path to APS XLSX.
    :return: pandas.DataFrame, curated stats key-values.
    """

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True)

    ws = wb["Stats"]
    current_section = "Files"
    records = []

    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value  # value
        b = ws.cell(row=r, column=2).value  # metric or section
        c = ws.cell(row=r, column=3).value  # details
        d = ws.cell(row=r, column=4).value  # comments

        b_norm = _norm_header(b)
        # Section headers in Stats are in column B when column A is empty.
        # Typically C/D are either blank or the "Details"/"Comments" labels for the first section.
        if a is None and b_norm:
            c_norm = _norm_header(c)
            d_norm = _norm_header(d)
            if (c is None and d is None) or (c_norm == "Details" and d_norm == "Comments"):
                current_section = b_norm
                continue

        # Skip blank metric labels
        if b is None or (isinstance(b, str) and b.strip() == ""):
            continue

        # Only emit metric rows when we have a value in column A
        if a is None:
            continue

        value_norm = a
        if isinstance(a, (datetime.datetime, datetime.date)):
            value_norm = _to_iso_date(a)
        elif isinstance(a, datetime.time):
            value_norm = _to_hms_time(a)
        if isinstance(value_norm, float) and value_norm.is_integer():
            value_norm = int(value_norm)


        rec = {
            "section": current_section,
            "metric": b_norm,
            "value": value_norm,
            "details": c if c is not None else "",
            "comments": d if d is not None else "",
            "source_sheet": "Stats",
            "source_row": r,
        }
        records.append(rec)

    df = pd.DataFrame(
        records,
        columns=[
            "section",
            "metric",
            "value",
            "details",
            "comments",
            "source_sheet",
            "source_row",
        ],
    )
    return df


def _build_aps_stats_referral_agreement_df(aps_xlsx_path):
    """Build the APS Stats referral agreement mini-table dataset.

    :param aps_xlsx_path: str, path to APS XLSX.
    :return: pandas.DataFrame, curated 2-row referral agreement table.
    """
    target = "Number of Tubes with FL Result"

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(aps_xlsx_path, data_only=True)

    ws = wb["Stats"]

    # Find anchor cell
    anchor_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == target:
                anchor_cell = cell
                break
        if anchor_cell is not None:
            break

    if anchor_cell is None:
        raise ValueError(f"Could not find anchor cell '{target}' in Stats sheet.")

    anchor_row = anchor_cell.row
    anchor_col = anchor_cell.column  # expected 6 (F)
    source_anchor = f"{openpyxl.utils.get_column_letter(anchor_col)}{anchor_row}"

    # Main header row spans F..L (6..12)
    main_headers = {}
    for col in range(6, 13):
        main_headers[col] = _norm_header(ws.cell(row=anchor_row, column=col).value)

    # Look for a breakdown header row that relabels I..L (9..12)
    breakdown_row = None
    for r in range(anchor_row + 1, min(anchor_row + 20, ws.max_row + 1)):
        v = _norm_header(ws.cell(row=r, column=9).value)  # I
        if v == "Ref/Cor Pos":
            breakdown_row = r
            break

    breakdown_headers = {}
    if breakdown_row is not None:
        for col in range(9, 13):
            breakdown_headers[col] = _norm_header(ws.cell(row=breakdown_row, column=col).value)

    # Find data rows for Positive and Inconclusive (column E == 5)
    def find_category_row(category):
        for r in range(anchor_row + 1, ws.max_row + 1):
            v = _norm_header(ws.cell(row=r, column=5).value)
            if v == category:
                return r
        return None

    pos_row = find_category_row("Positive")
    inc_row = find_category_row("Inconclusive")

    if pos_row is None or inc_row is None:
        raise ValueError("Could not find both Positive and Inconclusive rows in Stats mini-table.")

    # Decide header mode for each category:
    # - Positive: main agree/disagree
    # - Inconclusive: breakdown if there is a breakdown header between pos and inc
    inc_use_breakdown = breakdown_row is not None and pos_row < breakdown_row < inc_row

    out_cols = [
        "fl_result_category",
        "tubes_n",
        "cases_n",
        "with_ref_or_corresp_n",
        "agree_n",
        "agree_pct",
        "disagree_n",
        "disagree_pct",
        "ref_cor_pos_n",
        "incl_gt_pos_pct",
        "ref_cor_neg_n",
        "incl_gt_neg_pct",
        "source_sheet",
        "source_anchor",
    ]

    def read_row(category, row_num):
        tubes_n = ws.cell(row=row_num, column=6).value
        cases_n = ws.cell(row=row_num, column=7).value
        with_ref = ws.cell(row=row_num, column=8).value

        # Initialize all fields as empty
        rec = {k: "" for k in out_cols}
        rec["fl_result_category"] = category
        rec["tubes_n"] = tubes_n if tubes_n is not None else ""
        rec["cases_n"] = cases_n if cases_n is not None else ""
        rec["with_ref_or_corresp_n"] = with_ref if with_ref is not None else ""

        if category == "Positive":
            # Main header mode (agree/disagree)
            rec["agree_n"] = ws.cell(row=row_num, column=9).value or ""
            rec["agree_pct"] = ws.cell(row=row_num, column=10).value or ""
            rec["disagree_n"] = ws.cell(row=row_num, column=11).value or ""
            rec["disagree_pct"] = ws.cell(row=row_num, column=12).value or ""
        else:
            # Inconclusive: breakdown mode if present, else fall back to main
            if inc_use_breakdown:
                rec["ref_cor_pos_n"] = ws.cell(row=row_num, column=9).value or ""
                rec["incl_gt_pos_pct"] = ws.cell(row=row_num, column=10).value or ""
                rec["ref_cor_neg_n"] = ws.cell(row=row_num, column=11).value or ""
                rec["incl_gt_neg_pct"] = ws.cell(row=row_num, column=12).value or ""
            else:
                rec["agree_n"] = ws.cell(row=row_num, column=9).value or ""
                rec["agree_pct"] = ws.cell(row=row_num, column=10).value or ""
                rec["disagree_n"] = ws.cell(row=row_num, column=11).value or ""
                rec["disagree_pct"] = ws.cell(row=row_num, column=12).value or ""

        rec["source_sheet"] = "Stats"
        rec["source_anchor"] = source_anchor
        return rec

    records = [
        read_row("Positive", pos_row),
        read_row("Inconclusive", inc_row),
    ]

    df = pd.DataFrame(records, columns=out_cols)

    # Normalize numeric-ish fields (counts as Int64; percents as float)
    int_cols = ["tubes_n", "cases_n", "with_ref_or_corresp_n", "agree_n", "disagree_n", "ref_cor_pos_n", "ref_cor_neg_n"]
    float_cols = ["agree_pct", "disagree_pct", "incl_gt_pos_pct", "incl_gt_neg_pct"]
    for col in int_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    for col in float_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype(float)

    return df


def _build_weekly_summary_df(xlsx_path, sheet_name):
    """Build a weekly rollup dataset with stable headers.

    :param xlsx_path: str, path to XLSX.
    :param sheet_name: str, sheet name.
    :return: pandas.DataFrame, curated weekly summary.
    """
    df_raw = pd.read_excel(xlsx_path, sheet_name=sheet_name)

    # Drop unnamed / empty columns
    cols = []
    for c in df_raw.columns:
        c_norm = str(c)
        if c_norm.startswith("Unnamed:"):
            continue
        cols.append(c)
    df_raw = df_raw[cols]

    # Identify needed columns by normalized header
    col_map = {}
    for c in df_raw.columns:
        c_norm = _norm_header(c).lower()
        if c_norm == "week start date":
            col_map["week_start_date"] = c
        elif c_norm == "week end date":
            col_map["week_end_date"] = c
        elif c_norm.startswith("weekly participants"):
            col_map["participants_n"] = c
        elif c_norm.startswith("weekly tubes"):
            # Handles "Weekly Tubes Run 1379" and "Weekly Tubes 1379"
            col_map["tubes_n"] = c
        elif c_norm.startswith("weekly positive tubes"):
            col_map["positive_tubes_n"] = c
        elif c_norm.startswith("weekly inconclusive tubes"):
            col_map["inconclusive_tubes_n"] = c
        elif "% positive" in c_norm:
            col_map["pct_positive"] = c

    required = [
        "week_start_date",
        "week_end_date",
        "participants_n",
        "tubes_n",
        "positive_tubes_n",
        "inconclusive_tubes_n",
        "pct_positive",
    ]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise ValueError(f"Missing weekly columns in {sheet_name}: {missing}")

    df = df_raw[[col_map[k] for k in required]].copy()
    df.columns = required

    # Drop blank rows
    df = df[df["week_start_date"].notna()].copy()

    # Normalize dates
    df["week_start_date"] = df["week_start_date"].map(_to_iso_date)
    df["week_end_date"] = df["week_end_date"].map(_to_iso_date)

    # Counts
    for c in ["participants_n", "tubes_n", "positive_tubes_n", "inconclusive_tubes_n"]:
        df[c] = _coerce_int(df[c])

    df["pct_positive"] = _coerce_float(df["pct_positive"])

    # Compute pct_positive if missing
    missing_pct_mask = df["pct_positive"].isna() & (df["tubes_n"].fillna(0) > 0)
    if missing_pct_mask.any():
        df.loc[missing_pct_mask, "pct_positive"] = (
            df.loc[missing_pct_mask, "positive_tubes_n"].astype(float)
            / df.loc[missing_pct_mask, "tubes_n"].astype(float)
        )

    # Deterministic ordering (already in order, but enforce)
    df = df.sort_values(["week_start_date", "week_end_date"], kind="mergesort").reset_index(drop=True)
    return df


def _build_aps_pos_incl_cases_df(aps_xlsx_path):
    """Build APS Pos/Inconclusive cases dataset.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated pos/incl tube-level dataset.
    """
    df_raw = pd.read_excel(aps_xlsx_path, sheet_name="Pos and Incl")

    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()

    keep = {
        "TUBE ID": "tube_id",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "SPONSOR PERSON ID": "sponsor_id",
        "RESULT DATE": "result_date",
        "RESULT TIME": "result_time",
        "APP\nRESULT": "app_result",
        "FINAL RESULT": "final_result",
        "Result Correct Code": "result_correct_code",
        "Correction Notes": "correction_notes",
        "NUM SAMPLES (Pool Level)": "pool_level_n",
        "Sample Person1": "sample_person1",
        "Sample Person2": "sample_person2",
        "Sample Person3": "sample_person3",
        "Sample Person4": "sample_person4",
        "Case Culster ID": "case_cluster_id",
        "Index Tube - First FL Pos/Incl of Case?": "index_tube_first_case_flag",
        "Referral Test Result": "referral_test_result",
        "Referral Notes": "referral_notes",
        "3 ACCOUNTED FOR IN RTR?": "accounted_for_in_rtr_flag",
        "Re-test?": "retest_flag",
        "Pool Deconvolution Indiv Test?": "pool_deconvolution_indiv_test_flag",
        "Correspondence but no Referral Test": "correspondence_no_referral_flag",
        "Agreement with Referral or Correspondence?": "agreement_with_ref_or_correspondence_flag",
        "Known Positive (not Re-test)?": "known_positive_not_retest_flag",
        "FloodLAMP Result Eval": "floodlamp_result_eval",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    # Normalize date/time
    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["result_date"] = df["result_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["result_time"] = df["result_time"].map(_to_hms_time)

    # pool level as Int64 when possible
    df["pool_level_n"] = _coerce_int(df["pool_level_n"])

    # Fix spelling in output field: case_cluster_id already correct
    # Deterministic sort
    df = df.sort_values(["collection_date", "collection_time", "tube_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_aps_swabs_in_tubes_df(aps_xlsx_path):
    """Build APS swabs-in-tubes participant-result dataset from postDEL.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated swab-in-tube dataset.
    """
    df_raw = pd.read_excel(aps_xlsx_path, sheet_name="postDEL")

    # Filter rows with tube_id and participant_id
    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()
    df_raw = df_raw[df_raw["Participant ID"].notna()].copy()

    keep = {
        "TUBE ID": "tube_id",
        "Participant ID": "participant_id",
        "Sponsor ID": "sponsor_id",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "RESULT DATE": "result_date",
        "RESULT TIME": "result_time",
        "APP\nRESULT": "app_result",
        "FINAL RESULT": "final_result",
        "APP\nGROUP": "app_group",
        "FINAL\nGROUP": "final_group",
        "Basic Correct Code": "basic_correct_code",
        "Result Correct Code": "result_correct_code",
        "Correction Notes": "correction_notes",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    # Normalize
    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["result_date"] = df["result_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["result_time"] = df["result_time"].map(_to_hms_time)

    # IDs as strings (strip)
    for c in ["tube_id", "participant_id", "sponsor_id"]:
        df[c] = df[c].astype(str).str.strip()

    # Deterministic sort
    df = df.sort_values(["collection_date", "collection_time", "tube_id", "participant_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_aps_tubes_df(aps_xlsx_path):
    """Build APS tube-level dataset from APS Tubes sheet.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated tubes dataset.
    """
    df_raw = pd.read_excel(aps_xlsx_path, sheet_name="APS Tubes")
    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()

    keep = {
        "TUBE ID": "tube_id",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "SPONSOR PERSON ID": "sponsor_id",
        "NUM SAMPLES (Pool Level)": "pool_level_n",
        "APP\nRESULT": "app_result",
        "FINAL RESULT": "final_result",
        "RESULT DATE": "result_date",
        "RESULT TIME": "result_time",
        "Result Correct Code": "result_correct_code",
        "Correction Notes": "correction_notes",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["result_date"] = df["result_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["result_time"] = df["result_time"].map(_to_hms_time)

    df["pool_level_n"] = _coerce_int(df["pool_level_n"])

    # Deterministic sort
    df = df.sort_values(["collection_date", "collection_time", "tube_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_aps_daily_tubes_rollup_df(aps_xlsx_path):
    """Build APS daily tubes rollup dataset from APS Tubes by Date.

    :param aps_xlsx_path: str, APS XLSX path.
    :return: pandas.DataFrame, curated daily rollup.
    """
    df_raw = pd.read_excel(aps_xlsx_path, sheet_name="APS Tubes by Date", header=None)

    # Slice leftmost A..I (0..8)
    df_slice = df_raw.iloc[:, 0:9].copy()

    # Header row is first row
    header_row = df_slice.iloc[0].tolist()
    headers = [_strip_trailing_total(h) for h in header_row]
    headers = [_norm_header(h) for h in headers]
    df = df_slice.iloc[1:].copy()
    df.columns = headers

    # Drop blank rows based on collect date
    first_col = headers[0]
    df = df[df[first_col].notna()].copy()

    # Rename to stable schema
    rename_map = {
        "APS COLLECT DATE": "collect_date",
        "NUM TUBES": "tubes_n",
        "NEG": "neg_n",
        "POS": "pos_n",
        "INC": "inc_n",
        "NULL": "null_n",
        "INVL": "invl_n",
        "FAIL": "fail_n",
        "SUM RESULTS": "sum_results_n",
    }
    # Some sheets may keep newline; normalize keys
    norm_to_original = { _norm_header(c): c for c in df.columns }
    # Build an actual rename dict based on normalized headers
    rename_actual = {}
    for src_norm, dst in rename_map.items():
        if src_norm in norm_to_original:
            rename_actual[norm_to_original[src_norm]] = dst

    df = df.rename(columns=rename_actual).copy()

    required_cols = list(rename_map.values())
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"APS Tubes by Date missing expected columns after renaming: {missing}")

    df = df[required_cols].copy()

    # Normalize
    df["collect_date"] = df["collect_date"].map(_to_iso_date)
    for c in ["tubes_n", "neg_n", "pos_n", "inc_n", "null_n", "invl_n", "fail_n", "sum_results_n"]:
        df[c] = _coerce_int(df[c])

    df = df.sort_values(["collect_date"], kind="mergesort").reset_index(drop=True)
    return df


# -----------------------------
# RFR builders
# -----------------------------

def _build_rfr_all_tube_results_df(rfr_xlsx_path):
    """Build RFR all tube results dataset.

    :param rfr_xlsx_path: str, RFR XLSX path.
    :return: pandas.DataFrame, curated tube results with run linkage.
    """
    df_raw = pd.read_excel(rfr_xlsx_path, sheet_name="ALL Tube Results")
    df_raw = df_raw[df_raw["TUBE ID"].notna()].copy()

    keep = {
        "TUBE ID": "tube_id",
        "FINAL\nASSIGNED RUN ID": "final_assigned_run_id",
        "RUN DATE": "run_date",
        "COLLECTION DATE": "collection_date",
        "COLLECTION TIME": "collection_time",
        "FINAL RESULT": "final_result",
        "NUM SAMPLES (Pool Level)": "pool_level_n",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    df["run_date"] = df["run_date"].map(_to_iso_date)
    df["collection_date"] = df["collection_date"].map(_to_iso_date)
    df["collection_time"] = df["collection_time"].map(_to_hms_time)
    df["pool_level_n"] = _coerce_int(df["pool_level_n"])

    df = df.sort_values(["run_date", "final_assigned_run_id", "tube_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


# -----------------------------
# RTR builders
# -----------------------------

def _build_rtr_referral_tests_df(rtr_xlsx_path):
    """Build RTR referral tests dataset from Referral Test Data.

    :param rtr_xlsx_path: str, RTR XLSX path.
    :return: pandas.DataFrame, curated referral tests (row-level).
    """
    df_raw = pd.read_excel(rtr_xlsx_path, sheet_name="Referral Test Data")
    df_raw = df_raw[df_raw["REFERRAL TEST ID"].notna()].copy()

    keep = {
        "REFERRAL TEST ID": "referral_test_id",
        "PARTICIPANT ID": "participant_id",
        "CASE CLUSTER ID": "case_cluster_id",
        "REFERRAL TEST COLLECTION DATE": "referral_test_collection_date",
        "REFERRAL TEST COLLECTION TIME": "referral_test_collection_time",
        "REFERRAL TEST RESULT DATE": "referral_test_result_date",
        "REFERRAL TEST RESULT TIME": "referral_test_result_time",
        "REFERRAL TEST TYPE": "referral_test_type",
        "REFERRAL TEST MODEL": "referral_test_model",
        "REFERRAL TEST RESULT": "referral_test_result",
        "REFERRAL TEST CT (IF PCR)": "referral_test_ct",
        "SYMPTOMS?": "symptoms_flag",
        "VACCINATED?": "vaccinated_flag",
        "SOURCE NOTES": "source_notes",
        "INIT FL POS TUBE ID FROM CASE CLUSTER": "init_fl_pos_tube_id",
        "APPIVO COLLECTION DATE": "appivo_collection_date",
        "APPIVO COLLECTION TIME": "appivo_collection_time",
        "OTHER FL COLLECTION TIMESTAMP": "other_fl_collection_timestamp",
        "INIT POS": "init_pos",
        "INIT NEG": "init_neg",
        "INIT INC": "init_inc",
        "RERUN1 POS": "rerun1_pos",
        "RERUN1 NEG": "rerun1_neg",
        "RERUN1 INC": "rerun1_inc",
        "RERUN2 POS": "rerun2_pos",
        "RERUN2 NEG": "rerun2_neg",
        "RERUN2 INC": "rerun2_inc",
        "FL RESULT": "fl_result",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    # Normalize dates/times
    df["referral_test_collection_date"] = df["referral_test_collection_date"].map(_to_iso_date)
    df["referral_test_result_date"] = df["referral_test_result_date"].map(_to_iso_date)
    df["referral_test_collection_time"] = df["referral_test_collection_time"].map(_to_hms_time)
    df["referral_test_result_time"] = df["referral_test_result_time"].map(_to_hms_time)
    df["appivo_collection_date"] = df["appivo_collection_date"].map(_to_iso_date)
    df["appivo_collection_time"] = df["appivo_collection_time"].map(_to_hms_time)
    df["other_fl_collection_timestamp"] = df["other_fl_collection_timestamp"].map(_to_iso_datetime)

    # Numeric-ish fields
    if "referral_test_ct" in df.columns:
        df["referral_test_ct"] = pd.to_numeric(df["referral_test_ct"], errors="coerce")

    # Deterministic sort
    df = df.sort_values(
        ["participant_id", "referral_test_collection_date", "referral_test_collection_time", "referral_test_id"],
        kind="mergesort",
    ).reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


def _build_rtr_referral_tests_by_person_df(rtr_xlsx_path):
    """Build RTR referral tests by person dataset.

    :param rtr_xlsx_path: str, RTR XLSX path.
    :return: pandas.DataFrame, curated per-person referral summary.
    """
    df_raw = pd.read_excel(rtr_xlsx_path, sheet_name="Referral Tests by Person")
    df_raw = df_raw[df_raw["PARTICIPANT ID"].notna()].copy()

    keep = {
        "PARTICIPANT ID": "participant_id",
        "7 Num Sequential Referral Tests": "num_sequential_referral_tests",
        "Num FloodLAMP Results Pos or Incl": "num_floodlamp_results_pos_or_incl",
        "FloodLAMP Tube ID": "floodlamp_tube_id",
        "FloodLAMP Test Result": "floodlamp_test_result",
        "FloodLAMP Test Date": "floodlamp_test_date",
        "1st Referral Test Date": "first_referral_test_date",
        "Referral Overall Result": "referral_overall_result",
        "1st Referral Test Type": "first_referral_test_type",
        "1st Referral Test Result": "first_referral_test_result",
        "2nd Referral Test Type": "second_referral_test_type",
        "2nd Referral Test Result": "second_referral_test_result",
        "3rd Referral Test Type": "third_referral_test_type",
        "3rd Referral Test Result": "third_referral_test_result",
        "Antigen Neg with Other Positive": "antigen_neg_with_other_positive_flag",
        "Referral Eval": "referral_eval",
    }

    df = df_raw[list(keep.keys())].rename(columns=keep).copy()

    # Normalize dates
    df["floodlamp_test_date"] = df["floodlamp_test_date"].map(_to_iso_date)
    df["first_referral_test_date"] = df["first_referral_test_date"].map(_to_iso_date)

    # Coerce numeric-ish counts
    df["num_sequential_referral_tests"] = _coerce_int(df["num_sequential_referral_tests"])
    df["num_floodlamp_results_pos_or_incl"] = _coerce_int(df["num_floodlamp_results_pos_or_incl"])

    df = df.sort_values(["participant_id"], kind="mergesort").reset_index(drop=True)

    out_cols = list(keep.values())
    df = df[out_cols]
    return df


# -----------------------------
# Curate: write all outputs
# -----------------------------

def curate_site_public_csvs(aps_xlsx_path, rfr_xlsx_path, rtr_xlsx_path, output_folder_path, site_code):
    """Create all curated CSV outputs for a site from APS/RFR/RTR XLSX inputs.

    :param aps_xlsx_path: str, path to APS XLSX.
    :param rfr_xlsx_path: str, path to RFR XLSX.
    :param rtr_xlsx_path: str, path to RTR XLSX.
    :param output_folder_path: str, folder for CSV outputs.
    :param site_code: str, site code (e.g., ABRM).
    :return: dict, summary with written file paths and row counts.
    """
    output_folder_path = _ensure_folder(output_folder_path)
    site_code = (site_code or "").strip().upper()

    written = []
    rows_written = {}

    _print_section("CURATE: INPUTS")
    _print_kv("APS XLSX", aps_xlsx_path)
    _print_kv("RFR XLSX", rfr_xlsx_path)
    _print_kv("RTR XLSX", rtr_xlsx_path)
    _print_kv("Output folder", output_folder_path)
    _print_dash()

    # APS
    _print_section("CURATE: APS")
    aps_stats_kv_path = _join(output_folder_path, _out_name(site_code, "APS", "stats_key-values"))
    aps_stats_ref_path = _join(output_folder_path, _out_name(site_code, "APS", "stats_referral-agreement"))
    aps_weekly_path = _join(output_folder_path, _out_name(site_code, "APS", "weekly-summary"))
    aps_pos_incl_path = _join(output_folder_path, _out_name(site_code, "APS", "pos-incl-cases"))
    aps_swabs_path = _join(output_folder_path, _out_name(site_code, "APS", "swabs-in-tubes"))
    aps_tubes_path = _join(output_folder_path, _out_name(site_code, "APS", "tubes"))
    aps_daily_path = _join(output_folder_path, _out_name(site_code, "APS", "daily-tubes-rollup"))

    df_stats_kv = _build_aps_stats_key_values_df(aps_xlsx_path)
    _write_csv(df_stats_kv, aps_stats_kv_path)
    written.append(aps_stats_kv_path)
    rows_written[os.path.basename(aps_stats_kv_path)] = len(df_stats_kv)
    _print_kv("Wrote", f"{os.path.basename(aps_stats_kv_path)} ({len(df_stats_kv)} rows)")

    df_stats_ref = _build_aps_stats_referral_agreement_df(aps_xlsx_path)
    _write_csv(df_stats_ref, aps_stats_ref_path)
    written.append(aps_stats_ref_path)
    rows_written[os.path.basename(aps_stats_ref_path)] = len(df_stats_ref)
    _print_kv("Wrote", f"{os.path.basename(aps_stats_ref_path)} ({len(df_stats_ref)} rows)")

    df_weekly = _build_weekly_summary_df(aps_xlsx_path, "REF Weekly")
    _write_csv(df_weekly, aps_weekly_path)
    written.append(aps_weekly_path)
    rows_written[os.path.basename(aps_weekly_path)] = len(df_weekly)
    _print_kv("Wrote", f"{os.path.basename(aps_weekly_path)} ({len(df_weekly)} rows)")

    df_pos_incl = _build_aps_pos_incl_cases_df(aps_xlsx_path)
    _write_csv(df_pos_incl, aps_pos_incl_path)
    written.append(aps_pos_incl_path)
    rows_written[os.path.basename(aps_pos_incl_path)] = len(df_pos_incl)
    _print_kv("Wrote", f"{os.path.basename(aps_pos_incl_path)} ({len(df_pos_incl)} rows)")

    df_swabs = _build_aps_swabs_in_tubes_df(aps_xlsx_path)
    _write_csv(df_swabs, aps_swabs_path)
    written.append(aps_swabs_path)
    rows_written[os.path.basename(aps_swabs_path)] = len(df_swabs)
    _print_kv("Wrote", f"{os.path.basename(aps_swabs_path)} ({len(df_swabs)} rows)")

    df_tubes = _build_aps_tubes_df(aps_xlsx_path)
    _write_csv(df_tubes, aps_tubes_path)
    written.append(aps_tubes_path)
    rows_written[os.path.basename(aps_tubes_path)] = len(df_tubes)
    _print_kv("Wrote", f"{os.path.basename(aps_tubes_path)} ({len(df_tubes)} rows)")

    df_daily = _build_aps_daily_tubes_rollup_df(aps_xlsx_path)
    _write_csv(df_daily, aps_daily_path)
    written.append(aps_daily_path)
    rows_written[os.path.basename(aps_daily_path)] = len(df_daily)
    _print_kv("Wrote", f"{os.path.basename(aps_daily_path)} ({len(df_daily)} rows)")

    _print_dash()

    # RFR
    _print_section("CURATE: RFR")
    rfr_all_path = _join(output_folder_path, _out_name(site_code, "RFR", "all-tube-results"))
    rfr_weekly_path = _join(output_folder_path, _out_name(site_code, "RFR", "weekly-summary"))

    df_rfr_all = _build_rfr_all_tube_results_df(rfr_xlsx_path)
    _write_csv(df_rfr_all, rfr_all_path)
    written.append(rfr_all_path)
    rows_written[os.path.basename(rfr_all_path)] = len(df_rfr_all)
    _print_kv("Wrote", f"{os.path.basename(rfr_all_path)} ({len(df_rfr_all)} rows)")

    df_rfr_weekly = _build_weekly_summary_df(rfr_xlsx_path, "Weekly")
    _write_csv(df_rfr_weekly, rfr_weekly_path)
    written.append(rfr_weekly_path)
    rows_written[os.path.basename(rfr_weekly_path)] = len(df_rfr_weekly)
    _print_kv("Wrote", f"{os.path.basename(rfr_weekly_path)} ({len(df_rfr_weekly)} rows)")

    _print_dash()

    # RTR
    _print_section("CURATE: RTR")
    rtr_tests_path = _join(output_folder_path, _out_name(site_code, "RTR", "referral-tests"))
    rtr_by_person_path = _join(output_folder_path, _out_name(site_code, "RTR", "referral-tests-by-person"))

    df_rtr_tests = _build_rtr_referral_tests_df(rtr_xlsx_path)
    _write_csv(df_rtr_tests, rtr_tests_path)
    written.append(rtr_tests_path)
    rows_written[os.path.basename(rtr_tests_path)] = len(df_rtr_tests)
    _print_kv("Wrote", f"{os.path.basename(rtr_tests_path)} ({len(df_rtr_tests)} rows)")

    df_rtr_by_person = _build_rtr_referral_tests_by_person_df(rtr_xlsx_path)
    _write_csv(df_rtr_by_person, rtr_by_person_path)
    written.append(rtr_by_person_path)
    rows_written[os.path.basename(rtr_by_person_path)] = len(df_rtr_by_person)
    _print_kv("Wrote", f"{os.path.basename(rtr_by_person_path)} ({len(df_rtr_by_person)} rows)")

    _print_dash()

    summary = {
        "written": written,
        "rows_written": rows_written,
    }

    _print_section("CURATE: SUMMARY")
    _print_kv("Files written", len(written))
    _print_kv("Row counts", rows_written)
    _print_section("CURATE: DONE")
    return summary


# -----------------------------
# Verify: compare CSVs to XLSX
# -----------------------------

def verify_site_public_csvs(aps_xlsx_path, rfr_xlsx_path, rtr_xlsx_path, output_folder_path, site_code):
    """Verify that all curated CSV outputs match the XLSX inputs.

    Verification strategy:
    - Recompute the expected curated DataFrame from the XLSX.
    - Read the produced CSV as strings (no NA coercion).
    - Convert expected DataFrame to strings using the same formatting rules.
    - Assert exact match in shape, columns, and cell values.

    :param aps_xlsx_path: str, APS XLSX path.
    :param rfr_xlsx_path: str, RFR XLSX path.
    :param rtr_xlsx_path: str, RTR XLSX path.
    :param output_folder_path: str, folder containing the curated CSVs.
    :param site_code: str, site code.
    :return: dict, verification summary (pass/fail per file).
    """
    output_folder_path = os.path.abspath(output_folder_path)
    site_code = (site_code or "").strip().upper()

    _print_section("VERIFY: INPUTS")
    _print_kv("APS XLSX", aps_xlsx_path)
    _print_kv("RFR XLSX", rfr_xlsx_path)
    _print_kv("RTR XLSX", rtr_xlsx_path)
    _print_kv("CSV folder", output_folder_path)
    _print_dash()

    results = {}

    def verify_one(label, csv_filename, expected_df_builder):
        csv_path = _join(output_folder_path, csv_filename)
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"Missing expected CSV for verification: {csv_path}")

        expected_df = expected_df_builder()
        expected_str = _df_to_string_frame(expected_df)

        actual_str = _read_csv_as_strings(csv_path)

        # Ensure same columns + order
        if list(actual_str.columns) != list(expected_str.columns):
            raise AssertionError(
                f"[VERIFY FAIL] {label}: column mismatch\n"
                f"expected: {list(expected_str.columns)}\n"
                f"actual:   {list(actual_str.columns)}"
            )

        _assert_frames_equal(expected_str, actual_str, label)
        results[csv_filename] = "PASS"
        _print_kv("PASS", f"{label} → {csv_filename}")

    _print_section("VERIFY: APS")
    verify_one(
        "APS Stats key-values",
        _out_name(site_code, "APS", "stats_key-values"),
        lambda: _build_aps_stats_key_values_df(aps_xlsx_path),
    )
    verify_one(
        "APS Stats referral agreement",
        _out_name(site_code, "APS", "stats_referral-agreement"),
        lambda: _build_aps_stats_referral_agreement_df(aps_xlsx_path),
    )
    verify_one(
        "APS Weekly summary",
        _out_name(site_code, "APS", "weekly-summary"),
        lambda: _build_weekly_summary_df(aps_xlsx_path, "REF Weekly"),
    )
    verify_one(
        "APS Pos/Incl cases",
        _out_name(site_code, "APS", "pos-incl-cases"),
        lambda: _build_aps_pos_incl_cases_df(aps_xlsx_path),
    )
    verify_one(
        "APS Swabs in tubes",
        _out_name(site_code, "APS", "swabs-in-tubes"),
        lambda: _build_aps_swabs_in_tubes_df(aps_xlsx_path),
    )
    verify_one(
        "APS Tubes",
        _out_name(site_code, "APS", "tubes"),
        lambda: _build_aps_tubes_df(aps_xlsx_path),
    )
    verify_one(
        "APS Daily tubes rollup",
        _out_name(site_code, "APS", "daily-tubes-rollup"),
        lambda: _build_aps_daily_tubes_rollup_df(aps_xlsx_path),
    )
    _print_dash()

    _print_section("VERIFY: RFR")
    verify_one(
        "RFR All tube results",
        _out_name(site_code, "RFR", "all-tube-results"),
        lambda: _build_rfr_all_tube_results_df(rfr_xlsx_path),
    )
    verify_one(
        "RFR Weekly summary",
        _out_name(site_code, "RFR", "weekly-summary"),
        lambda: _build_weekly_summary_df(rfr_xlsx_path, "Weekly"),
    )
    _print_dash()

    _print_section("VERIFY: RTR")
    verify_one(
        "RTR Referral tests",
        _out_name(site_code, "RTR", "referral-tests"),
        lambda: _build_rtr_referral_tests_df(rtr_xlsx_path),
    )
    verify_one(
        "RTR Referral tests by person",
        _out_name(site_code, "RTR", "referral-tests-by-person"),
        lambda: _build_rtr_referral_tests_by_person_df(rtr_xlsx_path),
    )
    _print_dash()

    _print_section("VERIFY: SUMMARY")
    _print_kv("Verified files", len(results))
    _print_kv("Results", results)
    _print_section("VERIFY: DONE")
    return results


# -----------------------------
# Manifest (optional convenience)
# -----------------------------

def write_site_csv_manifest(output_folder_path, site_code, aps_xlsx_name, rfr_xlsx_name, rtr_xlsx_name):
    """Write a per-site CSV manifest describing the curated CSV outputs.

    This manifest is helpful in public archives because it tells users:
    - which curated CSVs exist
    - what sheet/tab they came from
    - the data grain and join keys

    :param output_folder_path: str, output folder.
    :param site_code: str, site code.
    :param aps_xlsx_name: str, APS workbook filename.
    :param rfr_xlsx_name: str, RFR workbook filename.
    :param rtr_xlsx_name: str, RTR workbook filename.
    :return: str, path to written manifest CSV.
    """
    output_folder_path = _ensure_folder(output_folder_path)
    site_code = (site_code or "").strip().upper()

    manifest_path = _join(output_folder_path, _manifest_name(site_code))

    rows = [
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "Stats",
            "output_csv": _out_name(site_code, "APS", "stats_key-values"),
            "data_level": "stats",
            "primary_key": "",
            "join_keys": "",
            "row_grain": "one row per metric",
            "notes": "Tidy key/value extraction from Stats A-D; sectioned by header rows in col B.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "Stats",
            "output_csv": _out_name(site_code, "APS", "stats_referral-agreement"),
            "data_level": "stats",
            "primary_key": "",
            "join_keys": "",
            "row_grain": "one row per FL result category",
            "notes": "Extract mini-table anchored at 'Number of Tubes with FL Result' (cols E-L).",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "REF Weekly",
            "output_csv": _out_name(site_code, "APS", "weekly-summary"),
            "data_level": "weekly_rollup",
            "primary_key": "",
            "join_keys": "week_start_date;week_end_date",
            "row_grain": "one row per week",
            "notes": "Weekly screening rollup (imported from RFR when present); stable headers.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "Pos and Incl",
            "output_csv": _out_name(site_code, "APS", "pos-incl-cases"),
            "data_level": "tube",
            "primary_key": "tube_id",
            "join_keys": "tube_id;case_cluster_id",
            "row_grain": "one row per pos/incl tube",
            "notes": "Case-centric pos/incl subset with pool membership + referral linkage.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "postDEL",
            "output_csv": _out_name(site_code, "APS", "swabs-in-tubes"),
            "data_level": "participant_result",
            "primary_key": "",
            "join_keys": "tube_id;participant_id",
            "row_grain": "one row per swab-in-tube row",
            "notes": "Trimmed participant-result table; drops QA scaffolding and blank PII columns.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "APS Tubes",
            "output_csv": _out_name(site_code, "APS", "tubes"),
            "data_level": "tube",
            "primary_key": "tube_id",
            "join_keys": "tube_id",
            "row_grain": "one row per tube",
            "notes": "Canonical tube table (Appivo tubes); pool level + timing + results + corrections.",
        },
        {
            "site_code": site_code,
            "source_workbook": aps_xlsx_name,
            "source_tab": "APS Tubes by Date",
            "output_csv": _out_name(site_code, "APS", "daily-tubes-rollup"),
            "data_level": "daily_rollup",
            "primary_key": "collect_date",
            "join_keys": "collect_date",
            "row_grain": "one row per collection date",
            "notes": "Exports only leftmost A-I rollup table; headers normalized to remove embedded totals.",
        },
        {
            "site_code": site_code,
            "source_workbook": rfr_xlsx_name,
            "source_tab": "ALL Tube Results",
            "output_csv": _out_name(site_code, "RFR", "all-tube-results"),
            "data_level": "tube",
            "primary_key": "tube_id",
            "join_keys": "tube_id;final_assigned_run_id",
            "row_grain": "one row per tube",
            "notes": "Run linkage + includes ARF tubes (not in Appivo).",
        },
        {
            "site_code": site_code,
            "source_workbook": rfr_xlsx_name,
            "source_tab": "Weekly",
            "output_csv": _out_name(site_code, "RFR", "weekly-summary"),
            "data_level": "weekly_rollup",
            "primary_key": "",
            "join_keys": "week_start_date;week_end_date",
            "row_grain": "one row per week",
            "notes": "Canonical RFR weekly rollup (when present); stable headers; drops Unnamed columns.",
        },
        {
            "site_code": site_code,
            "source_workbook": rtr_xlsx_name,
            "source_tab": "Referral Test Data",
            "output_csv": _out_name(site_code, "RTR", "referral-tests"),
            "data_level": "referral",
            "primary_key": "referral_test_id",
            "join_keys": "participant_id;case_cluster_id;init_fl_pos_tube_id",
            "row_grain": "one row per referral test",
            "notes": "Trimmed referral test rows; drops PII columns and internal matching helper columns.",
        },
        {
            "site_code": site_code,
            "source_workbook": rtr_xlsx_name,
            "source_tab": "Referral Tests by Person",
            "output_csv": _out_name(site_code, "RTR", "referral-tests-by-person"),
            "data_level": "referral",
            "primary_key": "participant_id",
            "join_keys": "participant_id",
            "row_grain": "one row per participant",
            "notes": "Per-person referral summary; drops PII and internal matching helper columns.",
        },
    ]

    fieldnames = [
        "site_code",
        "source_workbook",
        "source_tab",
        "output_csv",
        "data_level",
        "primary_key",
        "join_keys",
        "row_grain",
        "notes",
    ]

    with open(manifest_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    _print_section("MANIFEST")
    _print_kv("Wrote", manifest_path)
    _print_kv("Rows", len(rows))
    return manifest_path


# -----------------------------
# Manual run pattern (Cursor style)
# -----------------------------

def mrun_curate_and_verify():
    """Manual run stub (fill in paths, then run).

    :return: None
    """
    pass


if __name__ == "__main__":
   site_code = "ABRM"
   site_folder = "data/floodlamp/pilots/pilot-data/" + site_code + "_pilot-data"
   downloads_folder = site_folder + "/xlsx_downloads"
   aps_xlsx_path = os.path.join(downloads_folder, site_code + "_APS_deID_PUB.xlsx")
   rfr_xlsx_path = os.path.join(downloads_folder, site_code + "_RFR_deID_PUB.xlsx")
   rtr_xlsx_path = os.path.join(downloads_folder, site_code + "_RTR_deID_PUB.xlsx")
   output_folder_path = site_folder + "/curated_csvs_CHATGPT-52PRO"

   curate_site_public_csvs(
       aps_xlsx_path=aps_xlsx_path,
       rfr_xlsx_path=rfr_xlsx_path,
       rtr_xlsx_path=rtr_xlsx_path,
       output_folder_path=output_folder_path,
       site_code=site_code,
   )

   verify_site_public_csvs(
       aps_xlsx_path=aps_xlsx_path,
       rfr_xlsx_path=rfr_xlsx_path,
       rtr_xlsx_path=rtr_xlsx_path,
       output_folder_path=output_folder_path,
       site_code=site_code,
   )

   write_site_csv_manifest(
       output_folder_path=output_folder_path,
       site_code=site_code,
       aps_xlsx_name=os.path.basename(aps_xlsx_path),
       rfr_xlsx_name=os.path.basename(rfr_xlsx_path),
       rtr_xlsx_name=os.path.basename(rtr_xlsx_path),
   )