# ===== START OF FILE xlsx_curate.py =====
# Library of functions to extract and curate data from FloodLAMP XLSX files to CSVs

import os
import csv
import re
from datetime import datetime, date, time
from openpyxl import load_workbook
import pandas as pd

### HELPER FUNCTIONS

def _normalize_header(header):
    """
    Normalizes a header string by replacing newlines with spaces and trimming whitespace.

    :param header: string, the header text to normalize.
    :return normalized: string, the normalized header.
    """
    if header is None:
        return None
    if not isinstance(header, str):
        header = str(header)
    # Replace newlines with spaces
    header = header.replace('\n', ' ')
    # Collapse multiple spaces
    header = re.sub(r'\s+', ' ', header)
    # Trim whitespace
    header = header.strip()
    return header

def _strip_embedded_totals(header):
    """
    Strips embedded totals from header strings like 'Weekly Participants 2954' -> 'Weekly Participants'.

    :param header: string, the header text to process.
    :return cleaned: string, the header without embedded totals.
    """
    if header is None:
        return None
    # Remove trailing numbers (with possible space before)
    cleaned = re.sub(r'\s+\d+$', '', header)
    return cleaned

def _coerce_date(value):
    """
    Converts a date/datetime value to YYYY-MM-DD string format.

    :param value: various, the value to convert (datetime, date, string, etc.).
    :return date_str: string or None, the formatted date or None if conversion fails.
    """
    if value is None:
        return None
    # Handle pandas NaT
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try parsing common date formats
        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d']:
            try:
                return datetime.strptime(value, fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return value  # Return as-is if no format matches
    return str(value) if value else None

def _coerce_time(value):
    """
    Converts a time/datetime value to HH:MM:SS string format.

    :param value: various, the value to convert (datetime, time, string, etc.).
    :return time_str: string or None, the formatted time or None if conversion fails.
    """
    if value is None:
        return None
    # Handle pandas NaT
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.strftime('%H:%M:%S')
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        # Try parsing common time formats
        for fmt in ['%H:%M:%S', '%H:%M', '%I:%M:%S %p', '%I:%M %p']:
            try:
                return datetime.strptime(value, fmt).strftime('%H:%M:%S')
            except ValueError:
                continue
        return value
    return str(value) if value else None

def _coerce_datetime_iso(value):
    """
    Converts a datetime value to ISO format YYYY-MM-DDTHH:MM:SS.

    :param value: various, the value to convert.
    :return datetime_str: string or None, the formatted datetime or None if conversion fails.
    """
    if value is None:
        return None
    # Handle pandas NaT
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%dT%H:%M:%S')
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        return value
    return str(value) if value else None

def _coerce_int(value):
    """
    Converts a value to integer if possible.

    :param value: various, the value to convert.
    :return int_val: integer or original value.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        try:
            return int(float(value))
        except ValueError:
            return value
    return value

def _is_empty(value):
    """
    Checks if a value is empty (None, empty string, or NaN).

    :param value: various, the value to check.
    :return is_empty: boolean, True if the value is empty.
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == '':
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    return False

def _write_csv(rows, output_path, columns):
    """
    Writes rows to a CSV file with specified column order.

    :param rows: list of dicts, the data rows to write.
    :param output_path: string, path to the output CSV file.
    :param columns: list of strings, the column names in order.
    :return count: integer, number of rows written.
    """
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return len(rows)

def _get_cell_value(ws, row, col):
    """
    Gets a cell value from a worksheet by row and column indices.

    :param ws: worksheet, the openpyxl worksheet object.
    :param row: integer, the row number (1-indexed).
    :param col: integer, the column number (1-indexed).
    :return value: the cell value.
    """
    return ws.cell(row=row, column=col).value

def _find_header_map(df, rename_map):
    """
    Creates a mapping from original column names to renamed columns using flexible matching.

    :param df: dataframe, the pandas dataframe with original columns.
    :param rename_map: dict, mapping of original patterns to new names.
    :return col_map: dict, mapping of actual column names to new names.
    """
    col_map = {}
    normalized_cols = {_normalize_header(c): c for c in df.columns}
    
    for pattern, new_name in rename_map.items():
        pattern_normalized = _normalize_header(pattern)
        # Exact match
        if pattern_normalized in normalized_cols:
            col_map[normalized_cols[pattern_normalized]] = new_name
            continue
        # Strip totals and try again
        pattern_stripped = _strip_embedded_totals(pattern_normalized)
        for norm_col, orig_col in normalized_cols.items():
            norm_stripped = _strip_embedded_totals(norm_col)
            if pattern_stripped and norm_stripped and pattern_stripped.lower() == norm_stripped.lower():
                col_map[orig_col] = new_name
                break
            # Prefix match for columns like "Weekly Tubes Run 123"
            if pattern_stripped and norm_stripped and norm_stripped.lower().startswith(pattern_stripped.lower()):
                col_map[orig_col] = new_name
                break
    
    return col_map

### APS EXTRACTION FUNCTIONS

def curate_aps_stats(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts data from the APS Stats sheet and creates two curated CSVs.

    :param input_xlsx_path: string, path to the APS XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Load workbook with data_only=True to get computed values
    wb = load_workbook(input_xlsx_path, data_only=True)
    ws = wb['Stats']
    
    # PART A: Key-values extraction
    key_value_rows = []
    current_section = "Files"
    
    for row_num in range(1, 251):
        cell_a = _get_cell_value(ws, row_num, 1)  # Column A
        cell_b = _get_cell_value(ws, row_num, 2)  # Column B
        cell_c = _get_cell_value(ws, row_num, 3)  # Column C
        cell_d = _get_cell_value(ws, row_num, 4)  # Column D
        
        # Skip completely blank rows
        if _is_empty(cell_a) and _is_empty(cell_b):
            continue
        
        # Check for section header: A is empty, B has text
        if _is_empty(cell_a) and not _is_empty(cell_b):
            b_value = str(cell_b).strip()
            # Check if this looks like a section header (not a metric label)
            # Section headers are typically standalone text in B with empty A
            if b_value and not any(char.isdigit() for char in b_value[:2]):
                current_section = b_value
                continue
        
        # Check for metric row: B has label, A has value
        if not _is_empty(cell_b) and not _is_empty(cell_a):
            key_value_rows.append({
                'section': current_section,
                'metric': str(cell_b).strip() if cell_b else '',
                'value': cell_a if not isinstance(cell_a, str) else cell_a.strip(),
                'details': str(cell_c).strip() if cell_c and not _is_empty(cell_c) else '',
                'comments': str(cell_d).strip() if cell_d and not _is_empty(cell_d) else '',
                'source_sheet': 'Stats',
                'source_row': row_num
            })
    
    # Write key-values CSV
    kv_output_path = os.path.join(output_folder_path, f"{site_code}_APS_stats_key-values.csv")
    kv_columns = ['section', 'metric', 'value', 'details', 'comments', 'source_sheet', 'source_row']
    kv_count = _write_csv(key_value_rows, kv_output_path, kv_columns)
    result['written'].append(kv_output_path)
    result['rows_written']['key_values'] = kv_count
    
    # Get distinct sections in order
    sections_seen = []
    for row in key_value_rows:
        if row['section'] not in sections_seen:
            sections_seen.append(row['section'])
    
    # PART B: Referral agreement mini-table
    # Find anchor cell "Number of Tubes with FL Result"
    anchor_row = None
    anchor_col = None
    for row_num in range(1, 100):
        for col_num in range(1, 15):
            cell_val = _get_cell_value(ws, row_num, col_num)
            if cell_val and str(cell_val).strip() == "Number of Tubes with FL Result":
                anchor_row = row_num
                anchor_col = col_num
                break
        if anchor_row:
            break
    
    referral_rows = []
    if anchor_row:
        source_anchor = f"Stats!{chr(64+anchor_col)}{anchor_row}"
        
        # Header row is at anchor_row, columns F-L (6-12) typically
        # But we use anchor_col as F
        header_f = _normalize_header(str(_get_cell_value(ws, anchor_row, anchor_col) or ''))
        header_g = _normalize_header(str(_get_cell_value(ws, anchor_row, anchor_col+1) or ''))
        header_h = _normalize_header(str(_get_cell_value(ws, anchor_row, anchor_col+2) or ''))
        header_i = _normalize_header(str(_get_cell_value(ws, anchor_row, anchor_col+3) or ''))
        header_j = _normalize_header(str(_get_cell_value(ws, anchor_row, anchor_col+4) or ''))
        header_k = _normalize_header(str(_get_cell_value(ws, anchor_row, anchor_col+5) or ''))
        header_l = _normalize_header(str(_get_cell_value(ws, anchor_row, anchor_col+6) or ''))
        
        # Find Positive and Inconclusive rows
        for row_num in range(anchor_row + 1, anchor_row + 20):
            cell_e = _get_cell_value(ws, row_num, anchor_col - 1)  # Category column (E)
            if cell_e:
                category = str(cell_e).strip()
                if category in ["Positive", "Inconclusive"]:
                    row_data = {
                        'fl_result_category': category,
                        'tubes_n': _coerce_int(_get_cell_value(ws, row_num, anchor_col)),
                        'cases_n': _coerce_int(_get_cell_value(ws, row_num, anchor_col+1)),
                        'with_ref_or_corresp_n': _coerce_int(_get_cell_value(ws, row_num, anchor_col+2)),
                        'agree_n': None,
                        'agree_pct': None,
                        'disagree_n': None,
                        'disagree_pct': None,
                        'ref_cor_pos_n': None,
                        'incl_gt_pos_pct': None,
                        'ref_cor_neg_n': None,
                        'incl_gt_neg_pct': None,
                        'source_sheet': 'Stats',
                        'source_anchor': source_anchor
                    }
                    
                    # Check which header mode we're using for columns I-L
                    # Look at the header row or a potential sub-header row
                    val_i = _get_cell_value(ws, row_num, anchor_col+3)
                    val_j = _get_cell_value(ws, row_num, anchor_col+4)
                    val_k = _get_cell_value(ws, row_num, anchor_col+5)
                    val_l = _get_cell_value(ws, row_num, anchor_col+6)
                    
                    if category == "Positive":
                        # Use Agree/Disagree interpretation
                        row_data['agree_n'] = _coerce_int(val_i)
                        row_data['agree_pct'] = val_j
                        row_data['disagree_n'] = _coerce_int(val_k)
                        row_data['disagree_pct'] = val_l
                    else:  # Inconclusive
                        # Check for sub-header or use ref/cor interpretation
                        row_data['ref_cor_pos_n'] = _coerce_int(val_i)
                        row_data['incl_gt_pos_pct'] = val_j
                        row_data['ref_cor_neg_n'] = _coerce_int(val_k)
                        row_data['incl_gt_neg_pct'] = val_l
                    
                    referral_rows.append(row_data)
    else:
        result['warnings'].append("Could not find anchor 'Number of Tubes with FL Result' in Stats sheet")
    
    # Write referral agreement CSV
    ref_output_path = os.path.join(output_folder_path, f"{site_code}_APS_stats_referral-agreement.csv")
    ref_columns = ['fl_result_category', 'tubes_n', 'cases_n', 'with_ref_or_corresp_n',
                   'agree_n', 'agree_pct', 'disagree_n', 'disagree_pct',
                   'ref_cor_pos_n', 'incl_gt_pos_pct', 'ref_cor_neg_n', 'incl_gt_neg_pct',
                   'source_sheet', 'source_anchor']
    ref_count = _write_csv(referral_rows, ref_output_path, ref_columns)
    result['written'].append(ref_output_path)
    result['rows_written']['referral_agreement'] = ref_count
    
    wb.close()
    
    return result

def curate_aps_tubes(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts data from the APS Tubes sheet and creates a curated CSV.

    :param input_xlsx_path: string, path to the APS XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='APS Tubes')
    
    # Normalize column names
    df.columns = [_normalize_header(c) for c in df.columns]
    
    # Define column mapping
    rename_map = {
        'TUBE ID': 'tube_id',
        'COLLECTION DATE': 'collection_date',
        'COLLECTION TIME': 'collection_time',
        'SPONSOR PERSON ID': 'sponsor_id',
        'NUM SAMPLES (Pool Level)': 'pool_level_n',
        'APP RESULT': 'app_result',
        'FINAL RESULT': 'final_result',
        'RESULT DATE': 'result_date',
        'RESULT TIME': 'result_time',
        'Result Correct Code': 'result_correct_code',
        'Correction Notes': 'correction_notes'
    }
    
    # Drop columns
    drop_cols = ['SPONSOR FIRSTNAME', 'SPONSOR LASTNAME', 'NOTE', 
                 'Match Row in APS Data PHI - postDEL', 'First Row of Collection Date',
                 'Result Date Minus Collect Date']
    for col in drop_cols:
        norm_col = _normalize_header(col)
        if norm_col in df.columns:
            df = df.drop(columns=[norm_col])
    
    # Find and rename columns
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows where tube_id is not null
    if 'tube_id' in df.columns:
        df = df[df['tube_id'].notna() & (df['tube_id'] != '')]
    
    # Normalize dates and times
    if 'collection_date' in df.columns:
        df['collection_date'] = df['collection_date'].apply(_coerce_date)
    if 'collection_time' in df.columns:
        df['collection_time'] = df['collection_time'].apply(_coerce_time)
    if 'result_date' in df.columns:
        df['result_date'] = df['result_date'].apply(_coerce_date)
    if 'result_time' in df.columns:
        df['result_time'] = df['result_time'].apply(_coerce_time)
    if 'pool_level_n' in df.columns:
        df['pool_level_n'] = df['pool_level_n'].apply(_coerce_int)
    
    # Select and order columns
    output_columns = ['tube_id', 'collection_date', 'collection_time', 'sponsor_id',
                      'pool_level_n', 'app_result', 'final_result', 'result_date',
                      'result_time', 'result_correct_code', 'correction_notes']
    
    # Only include columns that exist
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    sort_cols = [c for c in ['collection_date', 'collection_time', 'tube_id'] if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols)
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_APS_tubes.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['tubes'] = len(df)
    
    return result

def curate_aps_daily_tubes_rollup(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts daily tubes rollup data from the APS Tubes by Date sheet.

    :param input_xlsx_path: string, path to the APS XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read without header to slice specific columns
    df = pd.read_excel(input_xlsx_path, sheet_name='APS Tubes by Date', header=None)
    
    # Slice columns A..I (0..8)
    df = df.iloc[:, 0:9]
    
    # Use first row as header
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    
    # Normalize headers and strip embedded totals
    df.columns = [_strip_embedded_totals(_normalize_header(c)) for c in df.columns]
    
    # Drop rows where first column is null
    first_col = df.columns[0]
    df = df[df[first_col].notna()]
    
    # Define column mapping
    rename_map = {
        'APS COLLECT DATE': 'collect_date',
        'NUM TUBES': 'tubes_n',
        'NEG': 'neg_n',
        'POS': 'pos_n',
        'INC': 'inc_n',
        'NULL': 'null_n',
        'INVL': 'invl_n',
        'FAIL': 'fail_n',
        'SUM RESULTS': 'sum_results_n'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Normalize dates
    if 'collect_date' in df.columns:
        df['collect_date'] = df['collect_date'].apply(_coerce_date)
        # Filter out non-date values
        df = df[df['collect_date'].notna()]
    
    # Convert count columns to int
    count_cols = ['tubes_n', 'neg_n', 'pos_n', 'inc_n', 'null_n', 'invl_n', 'fail_n', 'sum_results_n']
    for col in count_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_int)
    
    # Select and order columns
    output_columns = ['collect_date', 'tubes_n', 'neg_n', 'pos_n', 'inc_n', 
                      'null_n', 'invl_n', 'fail_n', 'sum_results_n']
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    if 'collect_date' in df.columns:
        df = df.sort_values(by='collect_date')
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_APS_daily-tubes-rollup.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['daily_rollup'] = len(df)
    
    # Check tubes_n == sum_results_n
    if 'tubes_n' in df.columns and 'sum_results_n' in df.columns:
        mismatches = df[df['tubes_n'] != df['sum_results_n']]
        if len(mismatches) > 0:
            result['warnings'].append(f"{len(mismatches)} rows where tubes_n != sum_results_n")
    
    return result

def curate_aps_pos_incl_cases(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts positive and inconclusive cases from the APS Pos and Incl sheet.

    :param input_xlsx_path: string, path to the APS XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='Pos and Incl')
    
    # Normalize column names
    df.columns = [_normalize_header(c) for c in df.columns]
    
    # Drop columns
    drop_cols = ['SPONSOR FIRSTNAME', 'SPONSOR LASTNAME', 'NOTE', 
                 'Match Row in APS Data PHI - postDEL', '0 Check']
    for col in drop_cols:
        norm_col = _normalize_header(col)
        if norm_col in df.columns:
            df = df.drop(columns=[norm_col])
    
    # Define column mapping
    rename_map = {
        'TUBE ID': 'tube_id',
        'COLLECTION DATE': 'collection_date',
        'COLLECTION TIME': 'collection_time',
        'SPONSOR PERSON ID': 'sponsor_id',
        'RESULT DATE': 'result_date',
        'RESULT TIME': 'result_time',
        'APP RESULT': 'app_result',
        'FINAL RESULT': 'final_result',
        'Result Correct Code': 'result_correct_code',
        'Correction Notes': 'correction_notes',
        'NUM SAMPLES (Pool Level)': 'pool_level_n',
        'Sample Person1': 'sample_person1',
        'Sample Person2': 'sample_person2',
        'Sample Person3': 'sample_person3',
        'Sample Person4': 'sample_person4',
        'Case Culster ID': 'case_cluster_id',
        'Index Tube - First FL Pos/Incl of Case?': 'index_tube_first_case_flag',
        'Referral Test Result': 'referral_test_result',
        'Referral Notes': 'referral_notes',
        '3 ACCOUNTED FOR IN RTR?': 'accounted_for_in_rtr_flag',
        'Re-test?': 'retest_flag',
        'Pool Deconvolution Indiv Test?': 'pool_deconvolution_indiv_test_flag',
        'Correspondence but no Referral Test': 'correspondence_no_referral_flag',
        'Agreement with Referral or Correspondence?': 'agreement_with_ref_or_correspondence_flag',
        'Known Positive (not Re-test)?': 'known_positive_not_retest_flag',
        'FloodLAMP Result Eval': 'floodlamp_result_eval'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows where tube_id is not null
    if 'tube_id' in df.columns:
        df = df[df['tube_id'].notna() & (df['tube_id'] != '')]
    
    # Normalize dates and times
    date_cols = ['collection_date', 'result_date']
    time_cols = ['collection_time', 'result_time']
    
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_date)
    for col in time_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_time)
    if 'pool_level_n' in df.columns:
        df['pool_level_n'] = df['pool_level_n'].apply(_coerce_int)
    
    # Select and order columns
    output_columns = ['tube_id', 'collection_date', 'collection_time', 'sponsor_id',
                      'result_date', 'result_time', 'app_result', 'final_result',
                      'result_correct_code', 'correction_notes', 'pool_level_n',
                      'sample_person1', 'sample_person2', 'sample_person3', 'sample_person4',
                      'case_cluster_id', 'index_tube_first_case_flag', 'referral_test_result',
                      'referral_notes', 'accounted_for_in_rtr_flag', 'retest_flag',
                      'pool_deconvolution_indiv_test_flag', 'correspondence_no_referral_flag',
                      'agreement_with_ref_or_correspondence_flag', 'known_positive_not_retest_flag',
                      'floodlamp_result_eval']
    
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    sort_cols = [c for c in ['collection_date', 'collection_time', 'tube_id'] if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols)
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_APS_pos-incl-cases.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['pos_incl_cases'] = len(df)
    
    return result

def curate_aps_swabs_in_tubes(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts swabs-in-tubes data from the APS postDEL sheet.

    :param input_xlsx_path: string, path to the APS XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='postDEL')
    
    # Normalize column names
    df.columns = [_normalize_header(c) for c in df.columns]
    
    # Drop PII and scaffolding columns
    drop_patterns = [
        'SPONSOR FIRSTNAME', 'SPONSOR LASTNAME', 'PARTICIPANT FIRSTNAME', 
        'PARTICIPANT LASTNAME', 'PARTICIPANT EMAIL', 'MINOR FIRSTNAME', 'MINOR LASTNAME',
        'Sponsor FullName', 'STATUS', 'Tube ID First Row?', 'Num Tube IDs', 'Add N Names',
        'NOTE FIRSTNAME', 'NOTE LASTNAME', 'Names Adjust', 'Note Name?', '>4 Samples',
        'Has Note?', 'Preserve Note', 'Initial Order', 'Added from APT?',
        'Minutes Col to Result', 'ERROR', 'Delete?', 'Delete Code', 'PARTCIP NAME COLS',
        'MINOR NAME COLS', 'NOTE NAME COLS', 'Note Name Extras', 'END'
    ]
    
    for col in list(df.columns):
        col_str = str(col)
        if any(pattern.lower() in col_str.lower() for pattern in drop_patterns):
            df = df.drop(columns=[col])
        # Drop numeric header columns like '0'
        if col_str.isdigit():
            df = df.drop(columns=[col])
    
    # Define column mapping
    rename_map = {
        'TUBE ID': 'tube_id',
        'Participant ID': 'participant_id',
        'Sponsor ID': 'sponsor_id',
        'COLLECTION DATE': 'collection_date',
        'COLLECTION TIME': 'collection_time',
        'RESULT DATE': 'result_date',
        'RESULT TIME': 'result_time',
        'APP RESULT': 'app_result',
        'FINAL RESULT': 'final_result',
        'APP GROUP': 'app_group',
        'FINAL GROUP': 'final_group',
        'Basic Correct Code': 'basic_correct_code',
        'Result Correct Code': 'result_correct_code',
        'Correction Notes': 'correction_notes'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows
    initial_count = len(df)
    if 'tube_id' in df.columns:
        df = df[df['tube_id'].notna() & (df['tube_id'] != '')]
    if 'participant_id' in df.columns:
        dropped_no_participant = len(df[df['participant_id'].isna() | (df['participant_id'] == '')])
        df = df[df['participant_id'].notna() & (df['participant_id'] != '')]
        if dropped_no_participant > 0:
            result['warnings'].append(f"Dropped {dropped_no_participant} rows with null participant_id")
    
    # Normalize dates and times
    date_cols = ['collection_date', 'result_date']
    time_cols = ['collection_time', 'result_time']
    
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_date)
    for col in time_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_time)
    
    # Strip whitespace from IDs
    for col in ['tube_id', 'participant_id', 'sponsor_id']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    
    # Select and order columns
    output_columns = ['tube_id', 'participant_id', 'sponsor_id', 'collection_date',
                      'collection_time', 'result_date', 'result_time', 'app_result',
                      'final_result', 'app_group', 'final_group', 'basic_correct_code',
                      'result_correct_code', 'correction_notes']
    
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    sort_cols = [c for c in ['collection_date', 'collection_time', 'tube_id', 'participant_id'] if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols)
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_APS_swabs-in-tubes.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['swabs_in_tubes'] = len(df)
    
    return result

def curate_aps_weekly_summary(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts weekly summary data from the APS REF Weekly sheet.

    :param input_xlsx_path: string, path to the APS XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='REF Weekly')
    
    # Normalize headers and strip embedded totals
    df.columns = [_strip_embedded_totals(_normalize_header(c)) for c in df.columns]
    
    # Drop unnamed columns
    df = df[[c for c in df.columns if not (c and str(c).startswith('Unnamed'))]]
    
    # Define column mapping with flexible matching
    rename_map = {
        'Week Start Date': 'week_start_date',
        'Week End Date': 'week_end_date',
        'Weekly Participants': 'participants_n',
        'Weekly Tubes': 'tubes_n',
        'Weekly Positive Tubes': 'positive_tubes_n',
        'Weekly Inconclusive Tubes': 'inconclusive_tubes_n',
        '% Positive': 'pct_positive'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows where week_start_date is not null
    if 'week_start_date' in df.columns:
        df = df[df['week_start_date'].notna()]
    
    # Normalize dates
    date_cols = ['week_start_date', 'week_end_date']
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_date)
            # Filter out non-date values
            df = df[df[col].notna()]
    
    # Convert count columns to int
    count_cols = ['participants_n', 'tubes_n', 'positive_tubes_n', 'inconclusive_tubes_n']
    for col in count_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_int)
    
    # Ensure pct_positive is float and compute if missing
    if 'pct_positive' in df.columns:
        df['pct_positive'] = pd.to_numeric(df['pct_positive'], errors='coerce')
    if 'tubes_n' in df.columns and 'positive_tubes_n' in df.columns:
        # Compute pct_positive where missing and tubes_n > 0
        mask = df['pct_positive'].isna() & (df['tubes_n'] > 0)
        df.loc[mask, 'pct_positive'] = df.loc[mask, 'positive_tubes_n'] / df.loc[mask, 'tubes_n']
    
    # Select and order columns
    output_columns = ['week_start_date', 'week_end_date', 'participants_n', 'tubes_n',
                      'positive_tubes_n', 'inconclusive_tubes_n', 'pct_positive']
    
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    if 'week_start_date' in df.columns:
        df = df.sort_values(by='week_start_date')
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_APS_weekly-summary.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['weekly_summary'] = len(df)
    
    return result

### RFR EXTRACTION FUNCTIONS

def curate_rfr_all_tube_results(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts all tube results data from the RFR ALL Tube Results sheet.

    :param input_xlsx_path: string, path to the RFR XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='ALL Tube Results')
    
    # Normalize column names
    df.columns = [_normalize_header(c) for c in df.columns]
    
    # Define column mapping
    rename_map = {
        'TUBE ID': 'tube_id',
        'FINAL ASSIGNED RUN ID': 'final_assigned_run_id',
        'RUN DATE': 'run_date',
        'COLLECTION DATE': 'collection_date',
        'COLLECTION TIME': 'collection_time',
        'FINAL RESULT': 'final_result',
        'NUM SAMPLES (Pool Level)': 'pool_level_n'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows where tube_id is not null
    if 'tube_id' in df.columns:
        df = df[df['tube_id'].notna() & (df['tube_id'] != '')]
    
    # Normalize dates and times
    date_cols = ['run_date', 'collection_date']
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_date)
    if 'collection_time' in df.columns:
        df['collection_time'] = df['collection_time'].apply(_coerce_time)
    if 'pool_level_n' in df.columns:
        df['pool_level_n'] = df['pool_level_n'].apply(_coerce_int)
    
    # Select and order columns
    output_columns = ['tube_id', 'final_assigned_run_id', 'run_date', 'collection_date',
                      'collection_time', 'final_result', 'pool_level_n']
    
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    sort_cols = [c for c in ['run_date', 'final_assigned_run_id', 'tube_id'] if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols)
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_RFR_all-tube-results.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['all_tube_results'] = len(df)
    
    return result

def curate_rfr_weekly_summary(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts weekly summary data from the RFR Weekly sheet.

    :param input_xlsx_path: string, path to the RFR XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='Weekly')
    
    # Normalize headers and strip embedded totals
    df.columns = [_strip_embedded_totals(_normalize_header(c)) for c in df.columns]
    
    # Drop unnamed/empty columns
    df = df[[c for c in df.columns if c and not str(c).startswith('Unnamed')]]
    
    # Define column mapping with flexible matching
    rename_map = {
        'Week Start Date': 'week_start_date',
        'Week End Date': 'week_end_date',
        'Weekly Participants': 'participants_n',
        'Weekly Tubes': 'tubes_n',
        'Weekly Tubes Run': 'tubes_n',  # Alternative naming
        'Weekly Positive Tubes': 'positive_tubes_n',
        'Weekly Inconclusive Tubes': 'inconclusive_tubes_n',
        '% Positive': 'pct_positive'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows where week_start_date is not null
    if 'week_start_date' in df.columns:
        df = df[df['week_start_date'].notna()]
    
    # Normalize dates
    date_cols = ['week_start_date', 'week_end_date']
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_date)
            df = df[df[col].notna()]
    
    # Convert count columns to int
    count_cols = ['participants_n', 'tubes_n', 'positive_tubes_n', 'inconclusive_tubes_n']
    for col in count_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_int)
    
    # Ensure pct_positive is float and compute if missing
    if 'pct_positive' in df.columns:
        df['pct_positive'] = pd.to_numeric(df['pct_positive'], errors='coerce')
    if 'tubes_n' in df.columns and 'positive_tubes_n' in df.columns:
        mask = df['pct_positive'].isna() & (df['tubes_n'] > 0)
        df.loc[mask, 'pct_positive'] = df.loc[mask, 'positive_tubes_n'] / df.loc[mask, 'tubes_n']
    
    # Select and order columns
    output_columns = ['week_start_date', 'week_end_date', 'participants_n', 'tubes_n',
                      'positive_tubes_n', 'inconclusive_tubes_n', 'pct_positive']
    
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    if 'week_start_date' in df.columns:
        df = df.sort_values(by='week_start_date')
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_RFR_weekly-summary.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['rfr_weekly_summary'] = len(df)
    
    return result

### RTR EXTRACTION FUNCTIONS

def curate_rtr_referral_tests(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts referral test data from the RTR Referral Test Data sheet.

    :param input_xlsx_path: string, path to the RTR XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='Referral Test Data')
    
    # Normalize column names
    df.columns = [_normalize_header(c) for c in df.columns]
    
    # Drop columns
    drop_patterns = [
        'Num Rows with Same Name in ALL Pos and Incl IMPORT',
        'Num Rows with Same Name in APS Person',
        'Num Rows with Same Name in Referall Tests',
        'PARTICIPANT FIRSTNAME', 'PARTICIPANT LASTNAME', 'PARTICIPANT FULLNAME',
        'NOTES'
    ]
    for col in list(df.columns):
        col_str = str(col)
        if any(pattern.lower() in col_str.lower() for pattern in drop_patterns):
            df = df.drop(columns=[col])
    
    # Define column mapping
    rename_map = {
        'CASE CLUSTER ID': 'case_cluster_id',
        'REFERRAL TEST ID': 'referral_test_id',
        'PARTICIPANT ID': 'participant_id',
        'REFERRAL TEST COLLECTION DATE': 'referral_test_collection_date',
        'REFERRAL TEST COLLECTION TIME': 'referral_test_collection_time',
        'REFERRAL TEST RESULT DATE': 'referral_test_result_date',
        'REFERRAL TEST RESULT TIME': 'referral_test_result_time',
        'REFERRAL TEST TYPE': 'referral_test_type',
        'REFERRAL TEST MODEL': 'referral_test_model',
        'REFERRAL TEST RESULT': 'referral_test_result',
        'REFERRAL TEST CT (IF PCR)': 'referral_test_ct',
        'SYMPTOMS?': 'symptoms_flag',
        'VACCINATED?': 'vaccinated_flag',
        'SOURCE NOTES': 'source_notes',
        'INIT FL POS TUBE ID FROM CASE CLUSTER': 'init_fl_pos_tube_id',
        'APPIVO COLLECTION DATE': 'appivo_collection_date',
        'APPIVO COLLECTION TIME': 'appivo_collection_time',
        'OTHER FL COLLECTION TIMESTAMP': 'other_fl_collection_timestamp',
        'INIT POS': 'init_pos',
        'INIT NEG': 'init_neg',
        'INIT INC': 'init_inc',
        'RERUN1 POS': 'rerun1_pos',
        'RERUN1 NEG': 'rerun1_neg',
        'RERUN1 INC': 'rerun1_inc',
        'RERUN2 POS': 'rerun2_pos',
        'RERUN2 NEG': 'rerun2_neg',
        'RERUN2 INC': 'rerun2_inc',
        'FL RESULT': 'fl_result'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows where referral_test_id is not null
    if 'referral_test_id' in df.columns:
        df = df[df['referral_test_id'].notna() & (df['referral_test_id'] != '')]
    
    # Normalize dates and times
    date_cols = ['referral_test_collection_date', 'referral_test_result_date', 'appivo_collection_date']
    time_cols = ['referral_test_collection_time', 'referral_test_result_time', 'appivo_collection_time']
    
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_date)
    for col in time_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_time)
    
    # Handle other_fl_collection_timestamp as ISO datetime
    if 'other_fl_collection_timestamp' in df.columns:
        df['other_fl_collection_timestamp'] = df['other_fl_collection_timestamp'].apply(_coerce_datetime_iso)
    
    # Select and order columns
    output_columns = ['referral_test_id', 'participant_id', 'case_cluster_id',
                      'referral_test_collection_date', 'referral_test_collection_time',
                      'referral_test_result_date', 'referral_test_result_time',
                      'referral_test_type', 'referral_test_model', 'referral_test_result',
                      'referral_test_ct', 'symptoms_flag', 'vaccinated_flag', 'source_notes',
                      'init_fl_pos_tube_id', 'appivo_collection_date', 'appivo_collection_time',
                      'other_fl_collection_timestamp', 'init_pos', 'init_neg', 'init_inc',
                      'rerun1_pos', 'rerun1_neg', 'rerun1_inc', 'rerun2_pos', 'rerun2_neg',
                      'rerun2_inc', 'fl_result']
    
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    sort_cols = [c for c in ['participant_id', 'referral_test_collection_date', 
                              'referral_test_collection_time', 'referral_test_id'] if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols)
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_RTR_referral-tests.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['referral_tests'] = len(df)
    
    return result

def curate_rtr_referral_tests_by_person(input_xlsx_path, output_folder_path, site_code):
    """
    Extracts referral tests by person data from the RTR Referral Tests by Person sheet.

    :param input_xlsx_path: string, path to the RTR XLSX file.
    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return result: dict, summary with written files, row counts, warnings, and errors.
    """
    result = {'written': [], 'rows_written': {}, 'warnings': [], 'errors': []}
    
    # Read the sheet
    df = pd.read_excel(input_xlsx_path, sheet_name='Referral Tests by Person')
    
    # Normalize column names
    df.columns = [_normalize_header(c) for c in df.columns]
    
    # Drop columns
    drop_patterns = [
        'PARTICIPANT FIRSTNAME', 'PARTICIPANT LASTNAME', 'PARTICIPANT FULLNAME',
        'Referral Match Row', 'FloodLAMP Match Row', 'Hardcoded Tube ID',
        'Same Tube ID', '3 Included in APS Pos and Incl?'
    ]
    for col in list(df.columns):
        col_str = str(col)
        if any(pattern.lower() in col_str.lower() for pattern in drop_patterns):
            df = df.drop(columns=[col])
        # Drop numeric header columns
        if col_str.isdigit():
            df = df.drop(columns=[col])
    
    # Define column mapping
    rename_map = {
        'PARTICIPANT ID': 'participant_id',
        '7 Num Sequential Referral Tests': 'num_sequential_referral_tests',
        'Num FloodLAMP Results Pos or Incl': 'num_floodlamp_results_pos_or_incl',
        'FloodLAMP Tube ID': 'floodlamp_tube_id',
        'FloodLAMP Test Result': 'floodlamp_test_result',
        'FloodLAMP Test Date': 'floodlamp_test_date',
        '1st Referral Test Date': 'first_referral_test_date',
        'Referral Overall Result': 'referral_overall_result',
        '1st Referral Test Type': 'first_referral_test_type',
        '1st Referral Test Result': 'first_referral_test_result',
        '2nd Referral Test Type': 'second_referral_test_type',
        '2nd Referral Test Result': 'second_referral_test_result',
        '3rd Referral Test Type': 'third_referral_test_type',
        '3rd Referral Test Result': 'third_referral_test_result',
        'Antigen Neg with Other Positive': 'antigen_neg_with_other_positive_flag',
        'Referral Eval': 'referral_eval'
    }
    
    col_map = _find_header_map(df, rename_map)
    df = df.rename(columns=col_map)
    
    # Filter rows where participant_id is not null
    if 'participant_id' in df.columns:
        df = df[df['participant_id'].notna() & (df['participant_id'] != '')]
    
    # Normalize dates
    date_cols = ['floodlamp_test_date', 'first_referral_test_date']
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_date)
    
    # Strip whitespace from IDs
    if 'participant_id' in df.columns:
        df['participant_id'] = df['participant_id'].astype(str).str.strip()
    
    # Select and order columns
    output_columns = ['participant_id', 'num_sequential_referral_tests', 'num_floodlamp_results_pos_or_incl',
                      'floodlamp_tube_id', 'floodlamp_test_result', 'floodlamp_test_date',
                      'first_referral_test_date', 'referral_overall_result', 'first_referral_test_type',
                      'first_referral_test_result', 'second_referral_test_type', 'second_referral_test_result',
                      'third_referral_test_type', 'third_referral_test_result',
                      'antigen_neg_with_other_positive_flag', 'referral_eval']
    
    available_cols = [c for c in output_columns if c in df.columns]
    df = df[available_cols]
    
    # Sort
    if 'participant_id' in df.columns:
        df = df.sort_values(by='participant_id')
    
    # Write output
    output_path = os.path.join(output_folder_path, f"{site_code}_RTR_referral-tests-by-person.csv")
    df.to_csv(output_path, index=False)
    result['written'].append(output_path)
    result['rows_written']['referral_tests_by_person'] = len(df)
    
    return result

### MANIFEST GENERATION

def generate_csv_manifest(output_folder_path, site_code):
    """
    Generates a CSV manifest file listing all curated outputs for a site.

    :param output_folder_path: string, path to the output folder.
    :param site_code: string, the site code (e.g., 'ABRM').
    :return output_path: string, path to the generated manifest file.
    """
    # Define manifest entries
    manifest_rows = [
        {'site_code': site_code, 'source_workbook': f"{site_code}_APS_deID_PUB.xlsx", 'source_tab': 'Stats', 'output_csv': f"{site_code}_APS_stats_key-values.csv", 'data_level': 'stats', 'primary_key': '', 'join_keys': '', 'row_grain': 'one row per metric', 'notes': 'tidy key/value extraction from Stats A-D, sectioned by headers'},
        {'site_code': site_code, 'source_workbook': f"{site_code}_APS_deID_PUB.xlsx", 'source_tab': 'Stats', 'output_csv': f"{site_code}_APS_stats_referral-agreement.csv", 'data_level': 'stats', 'primary_key': '', 'join_keys': '', 'row_grain': 'one row per FL result category (Positive/Inconclusive)', 'notes': 'mini-table extracted from Stats E-L anchored by "Number of Tubes with FL Result"'},
        {'site_code': site_code, 'source_workbook': f"{site_code}_APS_deID_PUB.xlsx", 'source_tab': 'REF Weekly', 'output_csv': f"{site_code}_APS_weekly-summary.csv", 'data_level': 'weekly_rollup', 'primary_key': '', 'join_keys': 'week_start_date;week_end_date', 'row_grain': 'one row per week', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_APS_deID_PUB.xlsx", 'source_tab': 'Pos and Incl', 'output_csv': f"{site_code}_APS_pos-incl-cases.csv", 'data_level': 'tube', 'primary_key': '', 'join_keys': 'tube_id;case_cluster_id', 'row_grain': 'one row per tube (pos/incl)', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_APS_deID_PUB.xlsx", 'source_tab': 'postDEL', 'output_csv': f"{site_code}_APS_swabs-in-tubes.csv", 'data_level': 'participant_result', 'primary_key': '', 'join_keys': '', 'row_grain': 'one row per swab-in-tube row', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_APS_deID_PUB.xlsx", 'source_tab': 'APS Tubes', 'output_csv': f"{site_code}_APS_tubes.csv", 'data_level': 'tube', 'primary_key': 'tube_id', 'join_keys': 'tube_id', 'row_grain': 'one row per tube', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_APS_deID_PUB.xlsx", 'source_tab': 'APS Tubes by Date', 'output_csv': f"{site_code}_APS_daily-tubes-rollup.csv", 'data_level': 'daily_rollup', 'primary_key': 'collect_date', 'join_keys': 'collect_date', 'row_grain': 'one row per collection date', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_RFR_deID_PUB.xlsx", 'source_tab': 'ALL Tube Results', 'output_csv': f"{site_code}_RFR_all-tube-results.csv", 'data_level': 'tube', 'primary_key': 'tube_id', 'join_keys': 'tube_id;final_assigned_run_id', 'row_grain': 'one row per tube', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_RFR_deID_PUB.xlsx", 'source_tab': 'Weekly', 'output_csv': f"{site_code}_RFR_weekly-summary.csv", 'data_level': 'weekly_rollup', 'primary_key': '', 'join_keys': 'week_start_date;week_end_date', 'row_grain': 'one row per week', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_RTR_deID_PUB.xlsx", 'source_tab': 'Referral Test Data', 'output_csv': f"{site_code}_RTR_referral-tests.csv", 'data_level': 'referral', 'primary_key': 'referral_test_id', 'join_keys': 'participant_id;case_cluster_id;init_fl_pos_tube_id', 'row_grain': 'one row per referral test', 'notes': ''},
        {'site_code': site_code, 'source_workbook': f"{site_code}_RTR_deID_PUB.xlsx", 'source_tab': 'Referral Tests by Person', 'output_csv': f"{site_code}_RTR_referral-tests-by-person.csv", 'data_level': 'referral', 'primary_key': 'participant_id', 'join_keys': 'participant_id', 'row_grain': 'one row per participant', 'notes': ''}
    ]
    
    # Write manifest
    output_path = os.path.join(output_folder_path, f"{site_code}_csv-manifest.csv")
    columns = ['site_code', 'source_workbook', 'source_tab', 'output_csv', 'data_level',
               'primary_key', 'join_keys', 'row_grain', 'notes']
    _write_csv(manifest_rows, output_path, columns)
    
    return output_path

### VERIFICATION FUNCTIONS

def _compare_values(csv_val, xlsx_val, col_name):
    """
    Compares a CSV value with an XLSX value, accounting for type coercion and date formatting.

    :param csv_val: the value from the CSV.
    :param xlsx_val: the value from the XLSX.
    :param col_name: string, the column name (for context).
    :return is_match: boolean, True if values are considered equivalent.
    """
    # Handle None/NaN/empty
    csv_empty = csv_val is None or (isinstance(csv_val, float) and pd.isna(csv_val)) or str(csv_val).strip() == ''
    xlsx_empty = xlsx_val is None or (isinstance(xlsx_val, float) and pd.isna(xlsx_val)) or str(xlsx_val).strip() == ''
    
    if csv_empty and xlsx_empty:
        return True
    if csv_empty != xlsx_empty:
        return False
    
    # Try exact string match
    if str(csv_val).strip() == str(xlsx_val).strip():
        return True
    
    # Try numeric comparison
    try:
        if float(csv_val) == float(xlsx_val):
            return True
    except (ValueError, TypeError):
        pass
    
    # Try date comparison
    csv_date = _coerce_date(csv_val)
    xlsx_date = _coerce_date(xlsx_val)
    if csv_date and xlsx_date and csv_date == xlsx_date:
        return True
    
    # Try time comparison
    csv_time = _coerce_time(csv_val)
    xlsx_time = _coerce_time(xlsx_val)
    if csv_time and xlsx_time and csv_time == xlsx_time:
        return True
    
    return False

def verify_csv_against_xlsx(csv_path, xlsx_path, sheet_name, key_column, csv_to_xlsx_column_map):
    """
    Verifies that a CSV file matches the source XLSX data.

    :param csv_path: string, path to the CSV file.
    :param xlsx_path: string, path to the source XLSX file.
    :param sheet_name: string, the sheet name in the XLSX.
    :param key_column: string, the column to use as a key for row matching.
    :param csv_to_xlsx_column_map: dict, mapping from CSV column names to XLSX column names.
    :return result: dict, with 'passed' boolean, 'total_rows', 'matched_rows', 'errors' list.
    """
    result = {'passed': True, 'total_rows': 0, 'matched_rows': 0, 'errors': []}
    
    # Read CSV
    csv_df = pd.read_csv(csv_path)
    result['total_rows'] = len(csv_df)
    
    # Read XLSX
    xlsx_df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    xlsx_df.columns = [_normalize_header(c) for c in xlsx_df.columns]
    
    # Check each row
    matched = 0
    for idx, csv_row in csv_df.iterrows():
        csv_key = csv_row.get(key_column)
        if pd.isna(csv_key):
            continue
        
        # Find matching row in XLSX
        xlsx_key_col = csv_to_xlsx_column_map.get(key_column, key_column)
        xlsx_matches = xlsx_df[xlsx_df[xlsx_key_col].astype(str) == str(csv_key)]
        
        if len(xlsx_matches) == 0:
            result['errors'].append(f"Row {idx}: No match found for {key_column}={csv_key}")
            result['passed'] = False
            continue
        
        xlsx_row = xlsx_matches.iloc[0]
        row_matched = True
        
        # Compare columns
        for csv_col, xlsx_col in csv_to_xlsx_column_map.items():
            if csv_col not in csv_df.columns:
                continue
            if xlsx_col not in xlsx_df.columns:
                continue
            
            csv_val = csv_row[csv_col]
            xlsx_val = xlsx_row[xlsx_col]
            
            if not _compare_values(csv_val, xlsx_val, csv_col):
                result['errors'].append(f"Row {idx}, col '{csv_col}': CSV='{csv_val}' != XLSX='{xlsx_val}'")
                result['passed'] = False
                row_matched = False
        
        if row_matched:
            matched += 1
    
    result['matched_rows'] = matched
    return result

def verify_all_csvs(output_folder_path, site_code, aps_xlsx_path, rfr_xlsx_path, rtr_xlsx_path):
    """
    Verifies all generated CSV files against their source XLSX files.

    :param output_folder_path: string, path to the folder containing CSV files.
    :param site_code: string, the site code (e.g., 'ABRM').
    :param aps_xlsx_path: string, path to the APS XLSX file.
    :param rfr_xlsx_path: string, path to the RFR XLSX file.
    :param rtr_xlsx_path: string, path to the RTR XLSX file.
    :return results: dict, mapping CSV names to verification results.
    """
    results = {}
    all_passed = True
    
    # Define verification configs
    verifications = [
        {'csv': f"{site_code}_APS_tubes.csv", 'xlsx': aps_xlsx_path, 'sheet': 'APS Tubes', 'key': 'tube_id', 'map': {'tube_id': 'TUBE ID', 'collection_date': 'COLLECTION DATE', 'final_result': 'FINAL RESULT'}},
        {'csv': f"{site_code}_APS_pos-incl-cases.csv", 'xlsx': aps_xlsx_path, 'sheet': 'Pos and Incl', 'key': 'tube_id', 'map': {'tube_id': 'TUBE ID', 'collection_date': 'COLLECTION DATE', 'final_result': 'FINAL RESULT'}},
        {'csv': f"{site_code}_RFR_all-tube-results.csv", 'xlsx': rfr_xlsx_path, 'sheet': 'ALL Tube Results', 'key': 'tube_id', 'map': {'tube_id': 'TUBE ID', 'run_date': 'RUN DATE', 'final_result': 'FINAL RESULT'}},
        {'csv': f"{site_code}_RTR_referral-tests.csv", 'xlsx': rtr_xlsx_path, 'sheet': 'Referral Test Data', 'key': 'referral_test_id', 'map': {'referral_test_id': 'REFERRAL TEST ID', 'participant_id': 'PARTICIPANT ID', 'referral_test_result': 'REFERRAL TEST RESULT'}},
        {'csv': f"{site_code}_RTR_referral-tests-by-person.csv", 'xlsx': rtr_xlsx_path, 'sheet': 'Referral Tests by Person', 'key': 'participant_id', 'map': {'participant_id': 'PARTICIPANT ID', 'referral_overall_result': 'Referral Overall Result'}}
    ]
    
    for v in verifications:
        csv_path = os.path.join(output_folder_path, v['csv'])
        if os.path.exists(csv_path) and os.path.exists(v['xlsx']):
            result = verify_csv_against_xlsx(csv_path, v['xlsx'], v['sheet'], v['key'], v['map'])
            results[v['csv']] = result
            if result['passed']:
                print(f"  PASSED: {v['csv']} ({result['matched_rows']}/{result['total_rows']} rows)")
            else:
                print(f"  FAILED: {v['csv']} ({result['matched_rows']}/{result['total_rows']} rows, {len(result['errors'])} errors)")
                all_passed = False
        else:
            print(f"  SKIPPED: {v['csv']} (file not found)")
    
    return results, all_passed

### MAIN PROCESSING FUNCTIONS

def process_site_xlsx_to_csv(site_code, xlsx_folder, output_folder):
    """
    Processes all XLSX files for a site and generates curated CSVs.

    :param site_code: string, the site code (e.g., 'ABRM').
    :param xlsx_folder: string, path to the folder containing XLSX files.
    :param output_folder: string, path to the output folder for CSVs.
    :return summary: dict, summary of all processing results.
    """
    summary = {'site_code': site_code, 'csvs_generated': [], 'total_rows': 0, 'errors': [], 'warnings': []}
    
    # Define file paths
    aps_xlsx = os.path.join(xlsx_folder, f"{site_code}_APS_deID_PUB.xlsx")
    rfr_xlsx = os.path.join(xlsx_folder, f"{site_code}_RFR_deID_PUB.xlsx")
    rtr_xlsx = os.path.join(xlsx_folder, f"{site_code}_RTR_deID_PUB.xlsx")
    
    # Print header
    print("\n" + "=" * 60)
    print(f"XLSX to CSV CURATION - Site: {site_code}")
    print("=" * 60)
    
    # Print input files
    print("\n=== INPUT FILES ===")
    input_files = [
        (aps_xlsx, "APS (All Participant Samples)"),
        (rfr_xlsx, "RFR (Run/Fluorescence Records)"),
        (rtr_xlsx, "RTR (Referral Test Records)")
    ]
    for fpath, label in input_files:
        exists = "FOUND" if os.path.exists(fpath) else "NOT FOUND"
        print(f"  [{exists}] {label}")
        print(f"           {fpath}")
    
    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)
    
    print(f"\n=== OUTPUT FOLDER ===")
    print(f"  {output_folder}")
    
    # Process files
    print(f"\n=== PROCESSING ===")
    
    # Process APS
    if os.path.exists(aps_xlsx):
        print(f"  Processing APS...")
        result = curate_aps_stats(aps_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['errors'].extend(result.get('errors', []))
        summary['warnings'].extend(result.get('warnings', []))
        for key, count in result['rows_written'].items():
            summary['total_rows'] += count
        
        result = curate_aps_tubes(aps_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('tubes', 0)
        
        result = curate_aps_daily_tubes_rollup(aps_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('daily_rollup', 0)
        summary['warnings'].extend(result.get('warnings', []))
        
        result = curate_aps_pos_incl_cases(aps_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('pos_incl_cases', 0)
        
        result = curate_aps_swabs_in_tubes(aps_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('swabs_in_tubes', 0)
        summary['warnings'].extend(result.get('warnings', []))
        
        result = curate_aps_weekly_summary(aps_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('weekly_summary', 0)
        print(f"    APS complete: 7 CSVs generated")
    else:
        summary['errors'].append(f"APS file not found: {aps_xlsx}")
        print(f"    APS SKIPPED (file not found)")
    
    # Process RFR
    if os.path.exists(rfr_xlsx):
        print(f"  Processing RFR...")
        result = curate_rfr_all_tube_results(rfr_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('all_tube_results', 0)
        
        result = curate_rfr_weekly_summary(rfr_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('rfr_weekly_summary', 0)
        print(f"    RFR complete: 2 CSVs generated")
    else:
        summary['errors'].append(f"RFR file not found: {rfr_xlsx}")
        print(f"    RFR SKIPPED (file not found)")
    
    # Process RTR
    if os.path.exists(rtr_xlsx):
        print(f"  Processing RTR...")
        result = curate_rtr_referral_tests(rtr_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('referral_tests', 0)
        
        result = curate_rtr_referral_tests_by_person(rtr_xlsx, output_folder, site_code)
        summary['csvs_generated'].extend(result['written'])
        summary['total_rows'] += result['rows_written'].get('referral_tests_by_person', 0)
        print(f"    RTR complete: 2 CSVs generated")
    else:
        summary['errors'].append(f"RTR file not found: {rtr_xlsx}")
        print(f"    RTR SKIPPED (file not found)")
    
    # Generate manifest
    manifest_path = generate_csv_manifest(output_folder, site_code)
    summary['csvs_generated'].append(manifest_path)
    print(f"  Manifest generated")
    
    # Print output files
    print(f"\n=== OUTPUT FILES ({len(summary['csvs_generated'])}) ===")
    for fpath in sorted(summary['csvs_generated']):
        print(f"  {os.path.basename(fpath)}")
    
    # Run verification
    print(f"\n=== VERIFICATION ===")
    verification_results, all_passed = verify_all_csvs(output_folder, site_code, aps_xlsx, rfr_xlsx, rtr_xlsx)
    summary['verification'] = verification_results
    
    # Print summary
    print("\n" + "=" * 60)
    if all_passed:
        print("RESULT: ALL VERIFICATIONS PASSED")
    else:
        print("RESULT: SOME VERIFICATIONS FAILED")
    print(f"Total CSVs: {len(summary['csvs_generated'])}")
    print(f"Total rows: {summary['total_rows']}")
    if summary['warnings']:
        print(f"Warnings: {len(summary['warnings'])}")
    if summary['errors']:
        print(f"Errors: {len(summary['errors'])}")
    print("=" * 60)
    
    return summary

def mrun_process_site_xlsx_to_csv():
    pass
#if __name__ == "__main__":
    site_code = "ABRM"
    xlsx_folder = "data/floodlamp/pilots/pilot-data/ABRM_pilot-data/ABRM_xlsx_downloads"
    output_folder = "data/floodlamp/pilots/pilot-data/ABRM_pilot-data/ABRM_curated_csvs"
    process_site_xlsx_to_csv(site_code, xlsx_folder, output_folder)

# ===== END OF FILE xlsx_curate.py =====
